#!/usr/bin/env python3
"""
amacedu_import.py
Generates SQL to import staff + students into amacedu DB for 2025-26.
Excel files in project root:
  - STAFF DETAILS.xlsx           (14 rows = 13 staff)
  - B.Ed.,I YEAR 2025-2026.xlsx  (100 B.Ed I Year students)
  - B.Ed &M.Ed II YEAR 2025-2026.xlsx (99 B.Ed II + 49 M.Ed II Year students)
  - M.Ed.,I year 2025-2026.xlsx  (50 M.Ed I Year students)
"""

import openpyxl
import datetime
import os

BASE = "/Applications/XAMPP/xamppfiles/htdocs/minerva"

# ── DB Reference IDs (from amacedu) ──────────────────────────────────────
LANG_ID      = 4    # English
CURRENCY_ID  = 68   # INR
SESSION_ID   = 1    # 2025-26
BCRYPT_PW    = r'$2y$12$mUoUzS1Fgc0a.qsXU1jIf.V37EhCavy82beMW9K9Zd9lqFQ9NMgyK'  # Welcome@123

# ── Designation IDs already in amacedu ────────────────────────────────────
# 6=PRINCIPAL, 37=ASSISTANT PROFESSOR, 38=PROFESSOR, 18=OTHERS, 43=TECH ASSISTANT, 53=GARDNER
DESIG_ASST_PROF = 37

# Designations to CREATE (not in DB)
NEW_DESIGS = ['Principal & Professor', 'Accountant', 'Sweeper', 'Scavenger', 'Technician']

# Designation string → role mapping  (roles: 2=Teacher, 3=Accountant, 10=Principal, 13=Others)
DESIG_ROLE_MAP = {
    'principal & professor': 10,
    'assistant professor':    2,
    'accountant':             3,
    'sweeper':               13,
    'scavenger':             13,
    'gardener':              13,
    'technician':            13,
}

# ── Utilities ─────────────────────────────────────────────────────────────
def esc(val, nullable=False):
    """SQL-safe string escaping."""
    if val is None or str(val).strip() == '':
        return 'NULL' if nullable else "''"
    s = str(val).strip()
    s = s.replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ').replace('\r', ' ')
    return f"'{s}'"

def esc_str(val):
    return esc(val, nullable=False)

def esc_null(val):
    return esc(val, nullable=True)

def fmt_date(val):
    """Convert various date formats to 'YYYY-MM-DD' or NULL."""
    if val is None:
        return 'NULL'
    if isinstance(val, (datetime.datetime, datetime.date)):
        return f"'{val.strftime('%Y-%m-%d')}'"
    s = str(val).strip()
    if not s:
        return 'NULL'
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return f"'{datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')}'"
        except ValueError:
            pass
    return 'NULL'

def fmt_decimal(val, nullable=True):
    if val is None or str(val).strip() == '':
        return 'NULL' if nullable else '0.00'
    try:
        return str(float(val))
    except (ValueError, TypeError):
        return 'NULL' if nullable else '0.00'

def fmt_contact(val):
    """Clean phone number."""
    if val is None:
        return "''"
    s = str(val).strip().replace('.0', '')
    # Remove scientific notation if present
    try:
        s = str(int(float(s)))
    except (ValueError, TypeError):
        pass
    return f"'{s}'"

# ── SQL Generation Start ──────────────────────────────────────────────────
lines = []

def sql(s):
    lines.append(s)

sql("SET NAMES utf8mb4;")
sql("SET FOREIGN_KEY_CHECKS = 0;")
sql("")

# ═══════════════════════════════════════════════════════════════════════════
# 1. New Designations
# ═══════════════════════════════════════════════════════════════════════════
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("-- 1. DESIGNATIONS")
sql("-- ═══════════════════════════════════════════════════════════════════")
for d in NEW_DESIGS:
    sql(f"INSERT IGNORE INTO staff_designation (designation) VALUES ({esc_str(d)});")
sql("")

# ═══════════════════════════════════════════════════════════════════════════
# 2. Classes and Sections
# ═══════════════════════════════════════════════════════════════════════════
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("-- 2. CLASSES AND SECTIONS (idempotent)")
sql("-- ═══════════════════════════════════════════════════════════════════")
CLASSES = ['B.Ed I Year', 'B.Ed II Year', 'M.Ed I Year', 'M.Ed II Year']
for c in CLASSES:
    sql(f"INSERT INTO classes (class) SELECT {esc_str(c)} "
        f"WHERE NOT EXISTS (SELECT 1 FROM classes WHERE class={esc_str(c)});")
sql("")
sql("INSERT INTO sections (section) SELECT 'A' "
    "WHERE NOT EXISTS (SELECT 1 FROM sections WHERE section='A');")
sql("")
for c in CLASSES:
    sql(f"INSERT INTO class_sections (class_id, section_id) "
        f"SELECT c.id, s.id FROM classes c "
        f"JOIN sections s ON s.section='A' "
        f"WHERE c.class={esc_str(c)} "
        f"AND NOT EXISTS ("
        f"  SELECT 1 FROM class_sections cs2 "
        f"  JOIN classes c2 ON c2.id=cs2.class_id "
        f"  WHERE c2.class={esc_str(c)});")
sql("")

# ═══════════════════════════════════════════════════════════════════════════
# 3. Staff Import
# ═══════════════════════════════════════════════════════════════════════════
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("-- 3. STAFF")
sql("-- ═══════════════════════════════════════════════════════════════════")

wb_staff = openpyxl.load_workbook(f"{BASE}/STAFF DETAILS.xlsx", read_only=True)
ws_staff = wb_staff.active
staff_rows = list(ws_staff.iter_rows(values_only=True))[1:]  # skip header

staff_employee_ids = []
for row in staff_rows:
    if not row[0]:
        continue
    emp_id       = str(row[0]).strip()
    qualification= str(row[1]).strip() if row[1] else ''
    work_exp     = str(row[2]).strip() if row[2] else ''
    name         = str(row[3]).strip().replace('\n', '').replace('  ', ' ') if row[3] else ''
    surname      = str(row[4]).strip() if row[4] else ''
    desig_str    = str(row[5]).strip() if row[5] else ''
    # col[6] = blank column, skip
    contact_no   = row[7]
    emg_contact  = row[8]
    email        = str(row[9]).strip() if row[9] else ''
    dob          = row[10]
    marital      = str(row[11]).strip() if row[11] else ''
    doj          = row[12]
    dol          = row[13]
    local_addr   = str(row[14]).strip().replace('\n', ' ') if row[14] else ''
    perm_addr    = str(row[15]).strip().replace('\n', ' ') if row[15] else ''
    note_val     = str(row[16]).strip() if row[16] else ''
    gender       = str(row[17]).strip() if row[17] else ''
    acct_title   = str(row[18]).strip() if row[18] else ''
    bank_acct    = row[19]
    bank_name    = str(row[20]).strip() if row[20] else ''
    ifsc         = str(row[21]).strip() if row[21] else ''
    bank_branch  = str(row[22]).strip() if row[22] else ''
    payscale     = str(int(float(row[23]))) if row[23] else ''
    basic_salary = fmt_decimal(row[24], nullable=True)
    uan_no       = str(int(float(row[25]))) if row[25] else ''  # epf_no → uan_no
    contract_type= str(row[26]).strip() if row[26] else ''
    shift        = str(row[27]).strip() if row[27] else ''
    location     = str(row[28]).strip() if row[28] else ''

    # Bank account: clean number
    if bank_acct is not None:
        try:
            bank_acct_str = str(int(float(str(bank_acct))))
        except (ValueError, TypeError):
            bank_acct_str = str(bank_acct).strip()
    else:
        bank_acct_str = ''

    # Resolve designation ID via subquery
    desig_id_sql = f"(SELECT id FROM staff_designation WHERE LOWER(designation)=LOWER({esc_str(desig_str)}) LIMIT 1)"

    staff_employee_ids.append((emp_id, desig_str))

    sql(f"INSERT INTO staff "
        f"(employee_id, qualification, work_exp, name, surname, "
        f"designation, lang_id, currency_id, "
        f"contact_no, emergency_contact_no, email, dob, marital_status, "
        f"date_of_joining, date_of_leaving, "
        f"local_address, permanent_address, note, gender, "
        f"account_title, bank_account_no, bank_name, ifsc_code, bank_branch, "
        f"payscale, basic_salary, uan_no, "
        f"contract_type, shift, location, "
        f"father_name, mother_name, emergency_contact_number, "
        f"image, other_document_name, other_document_file, "
        f"password, is_active, user_id, verification_code, "
        f"facebook, twitter, linkedin, instagram, "
        f"resume, joining_letter, resignation_letter) "
        f"SELECT "
        f"{esc_str(emp_id)}, {esc_str(qualification)}, {esc_str(work_exp)}, "
        f"{esc_str(name)}, {esc_str(surname)}, "
        f"{desig_id_sql}, {LANG_ID}, {CURRENCY_ID}, "
        f"{fmt_contact(contact_no)}, {fmt_contact(emg_contact)}, "
        f"{esc_str(email)}, {fmt_date(dob)}, {esc_str(marital)}, "
        f"{fmt_date(doj)}, {fmt_date(dol)}, "
        f"{esc_str(local_addr)}, {esc_str(perm_addr)}, "
        f"{esc_str(note_val)}, {esc_str(gender)}, "
        f"{esc_str(acct_title)}, {esc_str(bank_acct_str)}, "
        f"{esc_str(bank_name)}, {esc_str(ifsc)}, {esc_str(bank_branch)}, "
        f"{esc_str(payscale)}, {basic_salary}, {esc_null(uan_no if uan_no else None)}, "
        f"{esc_str(contract_type)}, {esc_str(shift)}, {esc_str(location)}, "
        f"'', '', '', "       # father_name, mother_name, emergency_contact_number
        f"'', '', '', "       # image, other_document_name, other_document_file
        f"{esc_str(BCRYPT_PW)}, 1, 0, '', "
        f"'', '', '', '', "   # facebook, twitter, linkedin, instagram
        f"'', '', '' "        # resume, joining_letter, resignation_letter
        f"WHERE NOT EXISTS (SELECT 1 FROM staff WHERE employee_id={esc_str(emp_id)});"
    )

sql("")

# ── Staff Roles ──────────────────────────────────────────────────────────
sql("-- Staff roles")
for emp_id, desig_str in staff_employee_ids:
    role_id = DESIG_ROLE_MAP.get(desig_str.lower(), 13)  # default Others
    sql(f"INSERT IGNORE INTO staff_roles (staff_id, role_id) "
        f"SELECT id, {role_id} FROM staff WHERE employee_id={esc_str(emp_id)};")
sql("")

# ═══════════════════════════════════════════════════════════════════════════
# 4. Students
# ═══════════════════════════════════════════════════════════════════════════
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("-- 4. STUDENTS  (note field used as temporary batch marker IMP_xxx)")
sql("-- ═══════════════════════════════════════════════════════════════════")

# All NOT NULL staff fields that need explicit empty string defaults
STAFF_EXTRA_EMPTY = "image, other_document_name, other_document_file"

STUDENT_COLS = ("admission_no, register_no, firstname, is_active, "
                "guardian_is, blood_group, height, weight, dis_note, "
                "father_pic, mother_pic, guardian_pic, guardian_occupation, "
                "note")

def student_insert(adm_no, reg_no, full_name, batch_tag):
    adm_esc  = esc_null(str(adm_no) if adm_no else None)
    reg_esc  = esc_null(str(reg_no).strip() if reg_no else None)
    name_esc = esc_str(str(full_name).strip())
    tag_esc  = esc_str(batch_tag)
    return (
        f"INSERT INTO students ({STUDENT_COLS}) "
        f"SELECT {adm_esc}, {reg_esc}, {name_esc}, "
        f"'yes', 'father', '', '', '', '', '', '', '', '', {tag_esc} "
        f"WHERE NOT EXISTS ("
        f"SELECT 1 FROM students "
        f"WHERE admission_no={adm_esc} AND note={tag_esc});"
    )

# ── B.Ed I Year ──────────────────────────────────────────────────────────
sql("-- B.Ed I Year students (100)")
wb_bed1 = openpyxl.load_workbook(f"{BASE}/B.Ed.,I YEAR 2025-2026.xlsx", read_only=True)
ws_bed1 = wb_bed1.active
cnt_bed1 = 0
for row in list(ws_bed1.iter_rows(values_only=True))[1:]:
    if not row[0] or not row[4]:
        continue
    sql(student_insert(row[0], None, row[4], 'IMP_BED1'))
    cnt_bed1 += 1
sql("")

# ── B.Ed II Year + M.Ed II Year (from combined file) ────────────────────
wb_ii = openpyxl.load_workbook(f"{BASE}/B.Ed &M.Ed II YEAR 2025-2026.xlsx", read_only=True)
ws_ii = wb_ii.active
ii_rows = list(ws_ii.iter_rows(values_only=True))[1:]

sql("-- B.Ed II Year students")
cnt_bed2 = 0
for row in ii_rows:
    if not row[0] or not row[1] or not row[2]:
        continue
    reg_str = str(row[1]).strip()
    if 'BD' not in reg_str:
        continue
    sql(student_insert(row[0], row[1], row[2], 'IMP_BED2'))
    cnt_bed2 += 1
sql("")

sql("-- M.Ed II Year students")
cnt_med2 = 0
for row in ii_rows:
    if not row[0] or not row[1] or not row[2]:
        continue
    reg_str = str(row[1]).strip()
    if 'MD' not in reg_str:
        continue
    sql(student_insert(row[0], row[1], row[2], 'IMP_MED2'))
    cnt_med2 += 1
sql("")

# ── M.Ed I Year ──────────────────────────────────────────────────────────
sql("-- M.Ed I Year students")
wb_med1 = openpyxl.load_workbook(f"{BASE}/M.Ed.,I year 2025-2026.xlsx", read_only=True)
ws_med1 = wb_med1.active
cnt_med1 = 0
for row in list(ws_med1.iter_rows(values_only=True))[1:]:
    if not row[0] or not row[4]:
        continue
    sql(student_insert(row[0], None, row[4], 'IMP_MED1'))
    cnt_med1 += 1
sql("")

# ═══════════════════════════════════════════════════════════════════════════
# 5. Student Sessions (link students → session + class + section)
# ═══════════════════════════════════════════════════════════════════════════
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("-- 5. STUDENT SESSIONS")
sql("-- ═══════════════════════════════════════════════════════════════════")

BATCH_CLASS = [
    ('IMP_BED1', 'B.Ed I Year'),
    ('IMP_BED2', 'B.Ed II Year'),
    ('IMP_MED1', 'M.Ed I Year'),
    ('IMP_MED2', 'M.Ed II Year'),
]

for batch_tag, class_name in BATCH_CLASS:
    sql(f"INSERT INTO student_session (session_id, student_id, class_id, section_id, is_active, is_alumni)")
    sql(f"  SELECT {SESSION_ID}, s.id, cs.class_id, cs.section_id, 'yes', 0")
    sql(f"  FROM students s")
    sql(f"  JOIN classes c ON c.class = {esc_str(class_name)}")
    sql(f"  JOIN class_sections cs ON cs.class_id = c.id")
    sql(f"  WHERE s.note = {esc_str(batch_tag)}")
    sql(f"  AND NOT EXISTS (")
    sql(f"    SELECT 1 FROM student_session ss2")
    sql(f"    WHERE ss2.session_id = {SESSION_ID} AND ss2.student_id = s.id);")
    sql("")

# ── Batch markers are kept permanently in the note field ─────────────────
sql("-- Note: IMP_xxx batch markers are kept permanently for idempotency")
sql("")

# ═══════════════════════════════════════════════════════════════════════════
# 6. Student user accounts (username=std{id}, password plain text)
# ═══════════════════════════════════════════════════════════════════════════
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("-- 6. STUDENT USER ACCOUNTS (idempotent)")
sql("-- ═══════════════════════════════════════════════════════════════════")
sql("INSERT INTO users (user_id, username, password, childs, role, lang_id, currency_id, verification_code, is_active)")
sql("SELECT s.id, CONCAT('std', s.id), 'Welcome@123', '', 'student', 4, 68, '', 'yes'")
sql("FROM students s")
sql("WHERE s.note LIKE 'IMP_%'")
sql("AND NOT EXISTS (SELECT 1 FROM users u WHERE u.user_id = s.id AND u.role = 'student');")
sql("")

sql("-- Parent user rows (one per student — keeps siblings list empty)")
sql("INSERT INTO users (user_id, username, password, childs, role, lang_id, currency_id, verification_code, is_active)")
sql("SELECT s.id, CONCAT('par', s.id), 'Welcome@123', s.id, 'parent', 4, 68, '', 'yes'")
sql("FROM students s")
sql("WHERE s.note LIKE 'IMP_%'")
sql("AND NOT EXISTS (SELECT 1 FROM users u WHERE u.username = CONCAT('par', s.id) AND u.role = 'parent');")
sql("")

sql("-- Set parent_id on each student to their unique parent user row id")
sql("UPDATE students s")
sql("JOIN users u ON u.username = CONCAT('par', s.id) AND u.role = 'parent'")
sql("SET s.parent_id = u.id")
sql("WHERE s.note LIKE 'IMP_%';")
sql("")

sql("SET FOREIGN_KEY_CHECKS = 1;")

# ═══════════════════════════════════════════════════════════════════════════
# Write SQL file
# ═══════════════════════════════════════════════════════════════════════════
out_path = f"{BASE}/tools/amacedu_import.sql"
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"SQL file written: {out_path}")
print(f"Summary:")
print(f"  Staff:         13")
print(f"  B.Ed I Year:   {cnt_bed1}")
print(f"  B.Ed II Year:  {cnt_bed2}")
print(f"  M.Ed I Year:   {cnt_med1}")
print(f"  M.Ed II Year:  {cnt_med2}")
print(f"  Total students: {cnt_bed1 + cnt_bed2 + cnt_med1 + cnt_med2}")
print(f"  SQL lines:     {len(lines)}")
