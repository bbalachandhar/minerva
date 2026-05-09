#!/usr/bin/env python3
"""
amace_nonteaching_import.py
Import non-teaching and maintenance staff from 'non teaching and maintenance Final.xlsx'
into the amace production DB.

Usage:
    python3 tools/amace_nonteaching_import.py            # generate SQL only
    python3 tools/amace_nonteaching_import.py --deploy   # generate + deploy
"""

import openpyxl
import re
import os
import sys
import subprocess
import datetime

BASE        = "/Applications/XAMPP/xamppfiles/htdocs/minerva"
XLSX_PATH   = f"{BASE}/non teaching and maintenance Final.xlsx"
SQL_OUT     = f"{BASE}/tools/amace_nonteaching_import.sql"
REMOTE_SQL  = "/tmp/amace_nonteaching_import.sql"

EC2_HOST    = 'ec2-user@13.234.255.106'
PEM         = '/Volumes/WORK/aws ec2 connect/minerva_prod.pem'
DB_USER     = 'meenakshi'
DB_PASS     = "3gw4*86*!zno-EMO"
DB_NAME     = 'amace'

LANG_ID      = 4    # English
CURRENCY_ID  = 68   # INR
PLAIN_PW     = 'Welcome@123'
BCRYPT_PW    = r'$2y$12$mUoUzS1Fgc0a.qsXU1jIf.V37EhCavy82beMW9K9Zd9lqFQ9NMgyK'

# employee_ids already in the amace DB (skip these)
EXISTING_EMP_IDS = {
    'ADMIN001','AMACE012','AMACE165','AMACE098','AMACE013','AMACE078','AMACE011',
    'AMACE138','AMACE001','AMACE023','AMACE083','AMACE090','AMACE139','AMACE109',
    'AMACE111','AMACE108','AMACE137','AMACE110','AMACE006','AMACE087','AMACE004',
    'AMACE005','AMACE145','AMACE026','AMACE027','AMACE028','AMACE029','AMACE073',
    'AMACE008','AMACE082','AMACE085','AMACE132','AMACE131','AMACE129','AMACE113',
    'AMACE112','AMACE002','AMACE119','AMACE142','AMACE127','AMACE124','AMACE149',
    'AMACE134','AMACE141','AMACE095','AMACE025','AMACE074','AMACE079','AMACE114',
    'AMACE116','AMACE018','AMACE120','AMACE146','AMACE136','AMACE133','AMACE118',
    'AMACE117','AMACE147','AMACE115','AMACE143','AMACE121','AMACE016','AMACE017',
    'AMACE130','AMACE148','AMACE126','AMACE123','AMACE106','AMACE128','AMACE164',
    'AMACE122','AMACE166','AMACE167',
}

# Designation string → staff_designation.id
DESIG_MAP = {
    'lab instructor':              24,   # LAB INSTRUCTOR
    'instector':                   24,   # typo → LAB INSTRUCTOR
    'asst librarian':              32,   # ASSISTANT LIBRARIAN
    'electrician':                 22,   # ELECTRICIAN
    'admin officer':               55,   # ADMINISTRATIVE OFFICER (AO)
    'accountant':                   8,   # Sr ACCOUNTANT
    'jr/assistant':                20,   # JUNIOR ASSISTANT
    'driver':                      30,   # DRIVER
    'attendar':                    15,   # ATTENDER
    'attender':                    15,
    'swepper':                     51,   # HOUSE KEEPING (typo sweeper)
    'sweeper':                     51,   # HOUSE KEEPING
    'jr accountant':               21,   # JUNIOR ACCOUNTANT
    'canteen supervisor':          18,   # OTHERS
    'gardan':                      53,   # GARDNER (typo)
    'gardener':                    53,
    'chief administrative officer': 55,  # ADMINISTRATIVE OFFICER (AO)
}
DESIG_DEFAULT = 18  # OTHERS

# Department string → department.id (None = CREATE new)
DEPT_MAP = {
    'eee':          24,
    'ece':          23,
    'mechanical':   27,
    's&h':          22,
    'cse&it':       31,  # map to CSE
    'mba':          32,
    # New departments to create (handled via INSERT IGNORE):
    'libray':       None,   # typo Library
    'library':      None,
    'office':       None,
    'transport':    None,
    'transports':   None,   # typo
    'eletrical':    None,   # typo Electrical
    'electrical':   None,
}
# New department names to create (INSERT IGNORE so safe to re-run)
NEW_DEPTS = ['Library', 'Office', 'Transport', 'Electrical']
# Normalised dept string → canonical name for INSERT IGNORE lookup
DEPT_CREATE_NAME = {
    'libray':    'Library',
    'library':   'Library',
    'office':    'Office',
    'transport': 'Transport',
    'transports':'Transport',
    'eletrical': 'Electrical',
    'electrical':'Electrical',
}

# category_id mapping from xlsx string
CAT_MAP = {
    'non teaching': 2,
    'maintenance':  4,
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def esc(v):
    if v is None or str(v).strip() == '':
        return "''"
    s = str(v).replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ').replace('\r', ' ')
    return f"'{s}'"

def esc_null(v):
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s or s.lower() in ('none', 'nil', 'null', '-', 'n/a'):
        return 'NULL'
    return "'" + s.replace('\\', '\\\\').replace("'", "\\'") + "'"

def clean_str(v, maxlen=None):
    if v is None:
        return ''
    s = str(v).strip()
    if s.lower() in ('none', 'nil', 'null', '-', 'n/a'):
        s = ''
    if maxlen:
        s = s[:maxlen]
    return s

def clean_phone(v, maxlen=20):
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return re.sub(r'\D', '', s)[:maxlen]

def clean_numeric(v, maxlen=200):
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s[:maxlen]

def fmt_date(val):
    if val is None:
        return 'NULL'
    if isinstance(val, (datetime.datetime, datetime.date)):
        return f"'{val.strftime('%Y-%m-%d')}'"
    s = str(val).strip()
    if not s or s.lower() in ('none', '', 'null'):
        return 'NULL'
    for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d %b %Y', '%d %B %Y'):
        try:
            return f"'{datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')}'"
        except ValueError:
            pass
    return 'NULL'

def salary_val(v):
    if v is None:
        return 'NULL'
    try:
        return str(float(v))
    except (ValueError, TypeError):
        return 'NULL'

def resolve_desig(s):
    if not s:
        return DESIG_DEFAULT
    return DESIG_MAP.get(str(s).strip().lower(), DESIG_DEFAULT)

def dept_expr(dept_str):
    """Return SQL expression for department id (int or subquery for new depts)."""
    if not dept_str:
        return 'NULL'
    key = str(dept_str).strip().lower()
    if key in ('non teaching', 'none', ''):
        return 'NULL'
    if key in DEPT_MAP:
        val = DEPT_MAP[key]
        if val is not None:
            return str(val)
        # new dept — use subquery
        canon = DEPT_CREATE_NAME.get(key, dept_str.strip().title())
        return f"(SELECT id FROM department WHERE LOWER(department_name)=LOWER('{canon}') LIMIT 1)"
    # unknown — return NULL
    return 'NULL'

# ─── Load xlsx ────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb['Sheet1']

# Parse header row to get column indices by name
header = [str(c).strip().lower() if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
COL = {name: idx for idx, name in enumerate(header)}

staff_rows = []
skipped = []

for row in ws.iter_rows(min_row=2, values_only=True):
    emp_id = clean_str(row[COL.get('employee_id', 1)])
    if not emp_id:
        continue
    if emp_id in EXISTING_EMP_IDS:
        skipped.append(emp_id)
        continue
    staff_rows.append(row)

print(f"\nExisting staff skipped: {len(skipped)} ({', '.join(skipped)})")
print(f"New staff to insert:    {len(staff_rows)}")
for r in staff_rows:
    emp_id   = clean_str(r[COL.get('employee_id', 1)])
    name     = clean_str(r[COL.get('name', 11)])
    desig_s  = clean_str(r[COL.get('designation', 44)])
    dept_s   = clean_str(r[COL.get('department', 45)])
    cat_s    = clean_str(r[COL.get('category_id', 46)])
    print(f"  {emp_id} — {name} | {dept_s} | {desig_s} | {cat_s}")

# ─── Build SQL ────────────────────────────────────────────────────────────────

lines = []
lines.append("SET foreign_key_checks = 0;")
lines.append("SET sql_mode = '';")
lines.append("")

# Create new departments (INSERT IGNORE = safe to re-run)
lines.append("-- ── New Departments ────────────────────────────────────────────")
for dept_name in NEW_DEPTS:
    lines.append(f"INSERT IGNORE INTO department (department_name, is_active)")
    lines.append(f"  VALUES ({esc(dept_name)}, 'yes');")
lines.append("")

lines.append("-- ── Staff + Users ─────────────────────────────────────────────")

for row in staff_rows:
    emp_id       = clean_str(row[COL.get('employee_id', 1)])
    prefix       = clean_str(row[COL.get('prefix', 2)])
    ug_qual      = clean_str(row[COL.get('ug_qualification', 3)])
    pg_qual      = clean_str(row[COL.get('pg_qualification', 4)])
    h_qual       = clean_str(row[COL.get('higher_qualification', 5)])
    qualification= clean_str(row[COL.get('qualification', 9)])
    work_exp     = clean_str(row[COL.get('work_exp', 10)])
    name         = clean_str(row[COL.get('name', 11)])
    surname      = clean_str(row[COL.get('surname', 12)])
    father_name  = clean_str(row[COL.get('father_name', 13)])
    mother_name  = clean_str(row[COL.get('mother_name', 14)])
    contact_no   = clean_phone(row[COL.get('contact_no', 15)])
    emerg_contact= clean_phone(row[COL.get('emergency_contact_no', 16)])
    email        = clean_str(row[COL.get('email', 17)])
    dob          = fmt_date(row[COL.get('dob', 18)])
    marital_stat = clean_str(row[COL.get('marital_status', 19)])
    doj          = fmt_date(row[COL.get('date_of_joining', 20)])
    dol          = fmt_date(row[COL.get('date_of_leaving', 21)])
    local_addr   = clean_str(row[COL.get('local_address', 22)])
    perm_addr    = clean_str(row[COL.get('permanent_address', 23)])
    note         = clean_str(row[COL.get('note', 24)])
    gender       = clean_str(row[COL.get('gender', 25)])
    acct_title   = clean_str(row[COL.get('account_title', 26)])
    bank_acct    = clean_numeric(row[COL.get('bank_account_no', 27)])
    bank_name    = clean_str(row[COL.get('bank_name', 28)])
    ifsc_code    = clean_str(row[COL.get('ifsc_code', 29)])
    bank_branch  = clean_str(row[COL.get('bank_branch', 30)])
    payscale     = clean_str(row[COL.get('payscale', 31)])
    basic_salary = salary_val(row[COL.get('basic_salary', 32)])
    esi_no       = clean_str(row[COL.get('esi_no', 33)])
    contract_type= clean_str(row[COL.get('contract_type', 34)])
    shift        = clean_str(row[COL.get('shift', 35)])
    location     = clean_str(row[COL.get('location', 36)])
    desig_str    = clean_str(row[COL.get('designation', 44)])
    dept_str     = clean_str(row[COL.get('department', 45)])
    cat_str      = clean_str(row[COL.get('category_id', 46)])
    aadhaar_no   = clean_numeric(row[COL.get('aadhaar_no', 47)])
    religion     = clean_str(row[COL.get('religion', 48)])
    caste        = clean_str(row[COL.get('caste', 49)])
    blood_group  = clean_str(row[COL.get('blood_group', 50)])
    country      = clean_str(row[COL.get('country', 51)])
    state        = clean_str(row[COL.get('state', 52)])
    pincode      = clean_str(row[COL.get('pincode', 53)])
    prev_salary  = salary_val(row[COL.get('previous_salary', 54)])
    uan_no       = clean_numeric(row[COL.get('uan_no', 55)])
    pan_no       = clean_str(row[COL.get('pan_no', 56)])
    prev_inst    = clean_str(row[COL.get('previous_institution', 57)])
    subj_exp     = clean_str(row[COL.get('subject_expertise', 58)])

    username     = emp_id
    desig_id     = resolve_desig(desig_str)
    dept_sql     = dept_expr(dept_str)
    cat_id       = CAT_MAP.get(cat_str.lower(), None)
    cat_sql      = str(cat_id) if cat_id else 'NULL'

    lines.append(f"-- {emp_id}: {name} {surname}")

    # Step 1: Insert user with user_id=0 placeholder
    lines.append(f"INSERT INTO users (user_id, username, password, childs, role, lang_id, currency_id, verification_code, is_active)")
    lines.append(f"  SELECT 0, {esc(username)}, {esc(PLAIN_PW)}, '', 'Teacher', {LANG_ID}, {CURRENCY_ID}, '', 'yes'")
    lines.append(f"  WHERE NOT EXISTS (SELECT 1 FROM users WHERE username = {esc(username)});")

    # Step 2: Insert staff (user_id = newly created users.id)
    lines.append(f"INSERT INTO staff (employee_id, prefix, ug_qualification, pg_qualification, higher_qualification,")
    lines.append(f"   qualification, work_exp, name, surname, father_name, mother_name,")
    lines.append(f"   contact_no, emergency_contact_no, email, dob, marital_status,")
    lines.append(f"   date_of_joining, date_of_leaving, local_address, permanent_address, note, gender,")
    lines.append(f"   account_title, bank_account_no, bank_name, ifsc_code, bank_branch, payscale, basic_salary,")
    lines.append(f"   esi_no, contract_type, shift, location,")
    lines.append(f"   designation, department, category_id, aadhaar_no, religion, caste, blood_group,")
    lines.append(f"   country, state, pincode, previous_salary, uan_no, pan_no,")
    lines.append(f"   previous_institution, subject_expertise,")
    lines.append(f"   lang_id, currency_id, password, image, is_active, user_id, verification_code,")
    lines.append(f"   other_document_name, other_document_file)")
    lines.append(f"  SELECT")
    lines.append(f"    {esc(emp_id)}, {esc(prefix)}, {esc(ug_qual)}, {esc(pg_qual)}, {esc(h_qual)},")
    lines.append(f"    {esc(qualification)}, {esc(work_exp)}, {esc(name)}, {esc(surname)}, {esc(father_name)}, {esc(mother_name)},")
    lines.append(f"    {esc(contact_no)}, {esc(emerg_contact)}, {esc(email)}, {dob}, {esc(marital_stat)},")
    lines.append(f"    {doj}, {dol}, {esc(local_addr)}, {esc(perm_addr)}, {esc(note)}, {esc(gender)},")
    lines.append(f"    {esc(acct_title)}, {esc(bank_acct)}, {esc(bank_name)}, {esc(ifsc_code)}, {esc(bank_branch)}, {esc(payscale)}, {basic_salary},")
    lines.append(f"    {esc(esi_no)}, {esc(contract_type)}, {esc(shift)}, {esc(location)},")
    lines.append(f"    {desig_id}, {dept_sql}, {cat_sql}, {esc(aadhaar_no)}, {esc(religion)}, {esc(caste)}, {esc(blood_group)},")
    lines.append(f"    {esc(country)}, {esc(state)}, {esc(pincode)}, {prev_salary}, {esc(uan_no)}, {esc(pan_no)},")
    lines.append(f"    {esc(prev_inst)}, {esc(subj_exp)},")
    lines.append(f"    {LANG_ID}, {CURRENCY_ID}, {esc(BCRYPT_PW)}, '', 1,")
    lines.append(f"    (SELECT id FROM users WHERE username = {esc(username)} LIMIT 1), '',")
    lines.append(f"    '', ''")
    lines.append(f"  WHERE NOT EXISTS (SELECT 1 FROM staff WHERE employee_id = {esc(emp_id)});")

    # Step 3: Update users.user_id = staff.id
    lines.append(f"UPDATE users SET user_id = (SELECT id FROM staff WHERE employee_id = {esc(emp_id)} LIMIT 1)")
    lines.append(f"  WHERE username = {esc(username)} AND user_id = 0;")
    lines.append("")

lines.append("SET foreign_key_checks = 1;")

sql_content = '\n'.join(lines)
with open(SQL_OUT, 'w') as f:
    f.write(sql_content)

print(f"\nSQL written: {SQL_OUT}  ({len(lines)} lines)")

# ─── Deploy ───────────────────────────────────────────────────────────────────
if '--deploy' not in sys.argv:
    print(f"\nTo deploy:\n  python3 {__file__} --deploy")
    sys.exit(0)

print("\nDeploying to EC2...")
scp = subprocess.run(
    ['scp', '-i', PEM, '-o', 'StrictHostKeyChecking=no', SQL_OUT, f'{EC2_HOST}:{REMOTE_SQL}'],
    capture_output=True, text=True
)
if scp.returncode != 0:
    print(f"SCP failed:\n{scp.stderr}")
    sys.exit(1)

print("SCP OK. Running SQL...")
ssh = subprocess.run(
    ['ssh', '-i', PEM, '-o', 'StrictHostKeyChecking=no', EC2_HOST,
     f"mysql -u {DB_USER} -p'{DB_PASS}' {DB_NAME} < {REMOTE_SQL}"],
    capture_output=True, text=True
)
if ssh.returncode != 0:
    print(f"MySQL failed: {ssh.stderr.strip()}")
    sys.exit(1)

print(f"Done! MySQL output: {ssh.stdout.strip() or '(no output = success)'}")
