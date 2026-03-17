from pathlib import Path
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path('/Applications/XAMPP/xamppfiles/htdocs/minerva')
ANALYSIS_JSON = ROOT / 'docs' / 'review_comments_jan2026_analysis.json'
OUT_XLSX = ROOT / 'Jan2026_review_comments_tabular_attachment.xlsx'


def style_header(ws, row_idx):
    fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def autosize(ws, min_width=12, max_width=55):
    widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(text) + 2)
    for col_idx, width in widths.items():
        col_letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, width))


def write_section_title(ws, row_idx, title):
    ws.cell(row=row_idx, column=1, value=title)
    ws.cell(row=row_idx, column=1).font = Font(bold=True, color='1F4E78')


def write_key_value_table(ws, start_row, title, rows):
    write_section_title(ws, start_row, title)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value='Metric')
    ws.cell(row=header_row, column=2, value='Value')
    style_header(ws, header_row)

    r = header_row + 1
    for key, value in rows:
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=value)
        r += 1
    return r + 1


def write_top_list(ws, start_row, title, items, label_key='label', count_key='count'):
    write_section_title(ws, start_row, title)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value='Rank')
    ws.cell(row=header_row, column=2, value='Item')
    ws.cell(row=header_row, column=3, value='Count')
    style_header(ws, header_row)

    r = header_row + 1
    for idx, item in enumerate(items, start=1):
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=item.get(label_key, ''))
        ws.cell(row=r, column=3, value=item.get(count_key, 0))
        r += 1
    return r + 1


def write_property_list(ws, start_row, title, items):
    write_section_title(ws, start_row, title)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value='Rank')
    ws.cell(row=header_row, column=2, value='Property Code')
    ws.cell(row=header_row, column=3, value='Property Name')
    ws.cell(row=header_row, column=4, value='Count')
    style_header(ws, header_row)

    r = header_row + 1
    for idx, item in enumerate(items, start=1):
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=item.get('property_code', ''))
        ws.cell(row=r, column=3, value=item.get('property_name', ''))
        ws.cell(row=r, column=4, value=item.get('count', 0))
        r += 1
    return r + 1


def write_sheet_summary(ws, start_row, sheet_name, sheet_data):
    r = write_key_value_table(ws, start_row, f'{sheet_name} Snapshot', [
        ('Sheet Name', sheet_name),
        ('Total Comments', sheet_data['total_rows']),
        ('Critical Comments', sheet_data['critical_count']),
        ('Non-Critical Comments', sheet_data['non_critical_count']),
        ('Critical %', f"{sheet_data['critical_pct']}%"),
        ('Unclassified Criticality', sheet_data['total_rows'] - sheet_data['critical_count'] - sheet_data['non_critical_count']),
    ])
    r = write_top_list(ws, r, f'{sheet_name} Top Categories', sheet_data['top_categories'][:10])
    r = write_top_list(ws, r, f'{sheet_name} Top Sub-categories', sheet_data['top_sub_categories'][:10])
    r = write_top_list(ws, r, f'{sheet_name} Top Controllers', sheet_data['top_controllers'][:10])
    r = write_top_list(ws, r, f'{sheet_name} Top Reviewers', sheet_data['top_reviewers'][:10])
    r = write_property_list(ws, r, f'{sheet_name} Top Properties by Volume', sheet_data['top_properties_by_volume'][:10])
    r = write_property_list(ws, r, f'{sheet_name} Critical Hotspots', sheet_data['critical_hotspots'][:10])
    return r


def write_three_column_comparison(ws, start_row, title, left_title, mid_title, right_title, rows):
    write_section_title(ws, start_row, title)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value='Metric')
    ws.cell(row=header_row, column=2, value=left_title)
    ws.cell(row=header_row, column=3, value=mid_title)
    ws.cell(row=header_row, column=4, value=right_title)
    style_header(ws, header_row)

    r = header_row + 1
    for metric, left_val, mid_val, right_val in rows:
        ws.cell(row=r, column=1, value=metric)
        ws.cell(row=r, column=2, value=left_val)
        ws.cell(row=r, column=3, value=mid_val)
        ws.cell(row=r, column=4, value=right_val)
        r += 1
    return r + 1


def write_ranked_comparison(ws, start_row, title, left_title, right_title, left_items, right_items, label_key='label', count_key='count'):
    write_section_title(ws, start_row, title)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value='Rank')
    ws.cell(row=header_row, column=2, value=f'{left_title} Item')
    ws.cell(row=header_row, column=3, value=f'{left_title} Count')
    ws.cell(row=header_row, column=4, value=f'{right_title} Item')
    ws.cell(row=header_row, column=5, value=f'{right_title} Count')
    style_header(ws, header_row)

    r = header_row + 1
    max_len = max(len(left_items), len(right_items))
    for idx in range(max_len):
        left = left_items[idx] if idx < len(left_items) else {}
        right = right_items[idx] if idx < len(right_items) else {}
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=left.get(label_key, left.get('property_name', '')))
        ws.cell(row=r, column=3, value=left.get(count_key, left.get('count', '')))
        ws.cell(row=r, column=4, value=right.get(label_key, right.get('property_name', '')))
        ws.cell(row=r, column=5, value=right.get(count_key, right.get('count', '')))
        r += 1
    return r + 1


def write_property_ranked_comparison(ws, start_row, title, left_title, right_title, left_items, right_items):
    write_section_title(ws, start_row, title)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value='Rank')
    ws.cell(row=header_row, column=2, value=f'{left_title} Property')
    ws.cell(row=header_row, column=3, value=f'{left_title} Count')
    ws.cell(row=header_row, column=4, value=f'{right_title} Property')
    ws.cell(row=header_row, column=5, value=f'{right_title} Count')
    style_header(ws, header_row)

    r = header_row + 1
    max_len = max(len(left_items), len(right_items))
    for idx in range(max_len):
        left = left_items[idx] if idx < len(left_items) else {}
        right = right_items[idx] if idx < len(right_items) else {}
        left_name = '' if not left else f"{left.get('property_code', '')} - {left.get('property_name', '')}"
        right_name = '' if not right else f"{right.get('property_code', '')} - {right.get('property_name', '')}"
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=left_name)
        ws.cell(row=r, column=3, value=left.get('count', ''))
        ws.cell(row=r, column=4, value=right_name)
        ws.cell(row=r, column=5, value=right.get('count', ''))
        r += 1
    return r + 1


def main():
    payload = json.loads(ANALYSIS_JSON.read_text(encoding='utf-8'))
    combined = payload['combined']
    sheets = payload['sheet_results']

    wb = Workbook()

    # Sheet 1: combined comparison summary
    ws1 = wb.active
    ws1.title = 'Summary_Comparison'

    r = 1
    r = write_key_value_table(ws1, r, 'Overall Snapshot', [
        ('Workbook', payload['workbook'].split('/')[-1]),
        ('Total Comments Reviewed', combined['total_rows']),
        ('Legacy Comments', combined['sheet_breakup'].get('Legacy', 0)),
        ('Asset west Comments', combined['sheet_breakup'].get('Asset west', 0)),
        ('Critical Comments', combined['critical_count']),
        ('Non-Critical Comments', combined['non_critical_count']),
        ('Critical %', f"{combined['critical_pct']}%"),
        ('Unclassified Criticality', combined['total_rows'] - combined['critical_count'] - combined['non_critical_count']),
    ])

    r = write_three_column_comparison(
        ws1,
        r,
        'Legacy vs Asset West Snapshot',
        'Legacy',
        'Asset west',
        'Combined',
        [
            ('Total Comments', sheets['Legacy']['total_rows'], sheets['Asset west']['total_rows'], combined['total_rows']),
            ('Critical Comments', sheets['Legacy']['critical_count'], sheets['Asset west']['critical_count'], combined['critical_count']),
            ('Non-Critical Comments', sheets['Legacy']['non_critical_count'], sheets['Asset west']['non_critical_count'], combined['non_critical_count']),
            ('Critical %', f"{sheets['Legacy']['critical_pct']}%", f"{sheets['Asset west']['critical_pct']}%", f"{combined['critical_pct']}%"),
            ('Unclassified Criticality', sheets['Legacy']['total_rows'] - sheets['Legacy']['critical_count'] - sheets['Legacy']['non_critical_count'], sheets['Asset west']['total_rows'] - sheets['Asset west']['critical_count'] - sheets['Asset west']['non_critical_count'], combined['total_rows'] - combined['critical_count'] - combined['non_critical_count']),
        ],
    )

    r = write_ranked_comparison(ws1, r, 'Top Categories: Legacy vs Asset West', 'Legacy', 'Asset west', sheets['Legacy']['top_categories'][:10], sheets['Asset west']['top_categories'][:10])
    r = write_ranked_comparison(ws1, r, 'Top Sub-categories: Legacy vs Asset West', 'Legacy', 'Asset west', sheets['Legacy']['top_sub_categories'][:10], sheets['Asset west']['top_sub_categories'][:10])
    r = write_property_ranked_comparison(ws1, r, 'Top Properties by Volume: Legacy vs Asset West', 'Legacy', 'Asset west', sheets['Legacy']['top_properties_by_volume'][:10], sheets['Asset west']['top_properties_by_volume'][:10])
    r = write_property_ranked_comparison(ws1, r, 'Critical Hotspots: Legacy vs Asset West', 'Legacy', 'Asset west', sheets['Legacy']['critical_hotspots'][:10], sheets['Asset west']['critical_hotspots'][:10])
    r = write_property_list(ws1, r, 'Combined Critical Hotspot Properties', combined['critical_hotspots'][:10])

    # Sheet 2: internal explanation in tabular action view
    ws2 = wb.create_sheet('Team_Explanation')
    r2 = 1
    r2 = write_key_value_table(ws2, r2, 'Key Interpretation', [
        ('Primary Volume Driver', 'Asset west sheet (75%+ of total comments)'),
        ('Primary Risk Driver', 'Asset west critical rate higher than Legacy'),
        ('Main Workstream', 'Dependency + Reclass + Accrual'),
        ('Fastest Risk Reduction', 'Focus on top 10 critical hotspot properties'),
        ('Data Quality Gap', 'Multiple category label variants; unclassified criticality present'),
    ])

    write_section_title(ws2, r2, 'Recommended Action Plan (Tabular)')
    ws2.cell(row=r2 + 1, column=1, value='Priority')
    ws2.cell(row=r2 + 1, column=2, value='Action')
    ws2.cell(row=r2 + 1, column=3, value='Owner')
    ws2.cell(row=r2 + 1, column=4, value='Target SLA')
    ws2.cell(row=r2 + 1, column=5, value='Expected Outcome')
    style_header(ws2, r2 + 1)

    actions = [
        ('P1', 'Close critical items on Y86, J69, F55, D29, J08 first', 'Property Leads', '48 hrs', 'Immediate critical backlog reduction'),
        ('P1', 'Dependency clearance with RM + Controller sync', 'Ops Lead', 'Daily', 'Remove external blockers faster'),
        ('P2', 'Dedicated Asset west critical triage lane', 'Review Team', 'Daily', 'Bring critical % below threshold'),
        ('P2', 'Standardize category/sub-category labels', 'QA + MIS', '3 days', 'Cleaner reporting and trend accuracy'),
        ('P2', 'Make criticality mandatory in tracker', 'Process Owner', '3 days', 'Remove unclassified rows from future cycles'),
    ]

    rr2 = r2 + 2
    for row in actions:
        for c, value in enumerate(row, start=1):
            ws2.cell(row=rr2, column=c, value=value)
        rr2 += 1
    r2 = rr2 + 1

    r2 = write_property_list(ws2, r2, 'Top Critical Hotspots for Allocation', combined['critical_hotspots'][:12])

    autosize(ws1)
    autosize(ws2)

    buffer = BytesIO()
    wb.save(buffer)
    OUT_XLSX.write_bytes(buffer.getvalue())
    print(str(OUT_XLSX))


if __name__ == '__main__':
    main()
