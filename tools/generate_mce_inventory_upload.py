"""
MCE Inventory Bulk Upload Generator
Reads: Stock Physical Verification-Completed file 120326.xlsx
Writes: MCE_Inventory_Upload_READY.xlsx

Sheets produced (in import order):
  1. README          — workflow instructions
  2. Categories      — ready to import (inventory_itemcategory)
  3. Supplier        — ready to import (inventory_itemsupplier) — one placeholder row
  4. Stores          — ready to import (inventory_itemstore) — one row per dept
  5. Items           — ready to import (inventory_item) — unique normalized items
  6. STOCK_FILL_PRICE — user fills purchase_price column, then import (inventory_itemstock)
"""

import re
from collections import defaultdict, OrderedDict
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter

SRC = '/Applications/XAMPP/xamppfiles/htdocs/minerva/Stock Physical Verification-Completed file 120326.xlsx'
OUT = '/Applications/XAMPP/xamppfiles/htdocs/minerva/tools/MCE_Inventory_Upload_READY.xlsx'

SUPPLIER = 'MCE Stock Verification 2026'
IMPORT_DATE = '2026-03-12'   # from filename 120326

# ── Category mapping ────────────────────────────────────────────────────────
# Map normalized item names (lowercase keywords) → category name
CATEGORY_MAP = OrderedDict([
    # IT & Computers
    ('cpu',           'IT & Computer Equipment'),
    ('keyboard',      'IT & Computer Equipment'),
    ('mouse',         'IT & Computer Equipment'),
    ('monitor',       'IT & Computer Equipment'),
    ('laptop',        'IT & Computer Equipment'),
    ('thin pc',       'IT & Computer Equipment'),
    ('server',        'IT & Computer Equipment'),
    ('switch',        'IT & Computer Equipment'),
    ('router',        'IT & Computer Equipment'),
    ('network',       'IT & Computer Equipment'),
    ('printer',       'IT & Computer Equipment'),
    ('scanner',       'IT & Computer Equipment'),
    ('webcam',        'IT & Computer Equipment'),
    ('pen drive',     'IT & Computer Equipment'),
    ('hard disk',     'IT & Computer Equipment'),
    ('ram',           'IT & Computer Equipment'),
    # Audio / Visual
    ('projector',     'Audio Visual Equipment'),
    ('led tv',        'Audio Visual Equipment'),
    ('smart board',   'Audio Visual Equipment'),
    ('interactive',   'Audio Visual Equipment'),
    ('lcd',           'Audio Visual Equipment'),
    ('speaker',       'Audio Visual Equipment'),
    ('microphone',    'Audio Visual Equipment'),
    ('amplifier',     'Audio Visual Equipment'),
    # Electrical & Power
    ('ups',           'Electrical & Power Equipment'),
    ('inverter',      'Electrical & Power Equipment'),
    ('battery',       'Electrical & Power Equipment'),
    ('stabilizer',    'Electrical & Power Equipment'),
    ('generator',     'Electrical & Power Equipment'),
    ('transformer',   'Electrical & Power Equipment'),
    ('electric panel','Electrical & Power Equipment'),
    ('mcb',           'Electrical & Power Equipment'),
    # Lighting & AC
    ('ceiling fan',   'Electrical Fixtures'),
    ('exhaust fan',   'Electrical Fixtures'),
    ('tube light',    'Electrical Fixtures'),
    ('led light',     'Electrical Fixtures'),
    ('led square',    'Electrical Fixtures'),
    ('cfl',           'Electrical Fixtures'),
    ('bulb',          'Electrical Fixtures'),
    ('light fitting', 'Electrical Fixtures'),
    ('air condition', 'Electrical Fixtures'),
    ('ac ',           'Electrical Fixtures'),
    # Furniture
    ('table',         'Furniture & Fixtures'),
    ('chair',         'Furniture & Fixtures'),
    ('stool',         'Furniture & Fixtures'),
    ('bench',         'Furniture & Fixtures'),
    ('rack',          'Furniture & Fixtures'),
    ('almirah',       'Furniture & Fixtures'),
    ('cupboard',      'Furniture & Fixtures'),
    ('locker',        'Furniture & Fixtures'),
    ('desk',          'Furniture & Fixtures'),
    ('sofa',          'Furniture & Fixtures'),
    ('berow',         'Furniture & Fixtures'),
    ('board',         'Furniture & Fixtures'),
    ('notice board',  'Furniture & Fixtures'),
    ('white board',   'Furniture & Fixtures'),
    ('black board',   'Furniture & Fixtures'),
    # Lab Instruments
    ('oscilloscope',  'Lab Instruments & Equipment'),
    ('multimeter',    'Lab Instruments & Equipment'),
    ('signal generator','Lab Instruments & Equipment'),
    ('function generator','Lab Instruments & Equipment'),
    ('power supply',  'Lab Instruments & Equipment'),
    ('spectrum',      'Lab Instruments & Equipment'),
    ('trainer kit',   'Lab Instruments & Equipment'),
    ('kit ',          'Lab Instruments & Equipment'),
    ('apparatus',     'Lab Instruments & Equipment'),
    ('instrument',    'Lab Instruments & Equipment'),
    ('balance',       'Lab Instruments & Equipment'),
    ('burette',       'Lab Instruments & Equipment'),
    ('pipette',       'Lab Instruments & Equipment'),
    ('beaker',        'Lab Instruments & Equipment'),
    ('flask',         'Lab Instruments & Equipment'),
    ('funnel',        'Lab Instruments & Equipment'),
    ('microscope',    'Lab Instruments & Equipment'),
    ('centrifuge',    'Lab Instruments & Equipment'),
    # Lab Consumables (chemicals, glassware, small components)
    ('resistor',      'Lab Consumables & Components'),
    ('resister',      'Lab Consumables & Components'),
    ('capacitor',     'Lab Consumables & Components'),
    ('transistor',    'Lab Consumables & Components'),
    ('transister',    'Lab Consumables & Components'),
    ('diode',         'Lab Consumables & Components'),
    ('ic ',           'Lab Consumables & Components'),
    ('test tube',     'Lab Consumables & Components'),
    ('reagent',       'Lab Consumables & Components'),
    ('chemical',      'Lab Consumables & Components'),
    ('glass',         'Lab Consumables & Components'),
    ('crucible',      'Lab Consumables & Components'),
    ('wire ',         'Lab Consumables & Components'),
    ('cable',         'Lab Consumables & Components'),
    # Office Equipment
    ('telephone',     'Office Equipment'),
    ('intercom',      'Office Equipment'),
    ('fax',           'Office Equipment'),
    ('calculator',    'Office Equipment'),
    ('shredder',      'Office Equipment'),
    ('photocopier',   'Office Equipment'),
    ('laminator',     'Office Equipment'),
    # Safety & Facilities
    ('fire extinguisher', 'Safety & Facilities'),
    ('cctv',          'Safety & Facilities'),
    ('camera',        'Safety & Facilities'),
    ('biometric',     'Safety & Facilities'),
    ('access control','Safety & Facilities'),
    ('first aid',     'Safety & Facilities'),
    # Library
    ('book',          'Library Resources'),
    ('library',       'Library Resources'),
    # Sports / Gym
    ('gym',           'Sports & Gymnasium'),
    ('treadmill',     'Sports & Gymnasium'),
    ('dumbbell',      'Sports & Gymnasium'),
    ('sports',        'Sports & Gymnasium'),
])

# ── Software license items to add manually (not from verification file) ────
# (item_name, unit, description, license_valid_from, license_valid_till)
SOFTWARE_ITEMS = [
    ('Windows OS License',          'Nos', 'Windows 10/11 Pro OEM License',       '2020-01-01', ''),
    ('MS Office License',           'Nos', 'Microsoft Office 365 / 2021',          '2024-04-01', '2025-03-31'),
    ('Antivirus License',           'Nos', 'Kaspersky / Quick Heal endpoint',       '2025-04-01', '2026-03-31'),
    ('AutoCAD License',             'Nos', 'AutoCAD (B.Arch / CIVIL / MECH)',       '2024-04-01', '2025-03-31'),
    ('MATLAB License',              'Nos', 'MATLAB + Simulink (EEE/ECE/EIE)',       '2024-04-01', '2025-03-31'),
    ('Tally Prime License',         'Nos', 'Tally Prime for accounts/admin',        '2024-04-01', '2025-03-31'),
    ('Adobe Creative Suite',        'Nos', 'Adobe CC (B.Arch / PD Cell)',           '2024-04-01', '2025-03-31'),
    ('SolidWorks License',          'Nos', 'SolidWorks CAD (MECH)',                 '2024-04-01', '2025-03-31'),
    ('ANSYS License',               'Nos', 'ANSYS simulation software',             '2024-04-01', '2025-03-31'),
    ('NS2/NS3 Simulator',           'Nos', 'Network simulator (CSC/IT)',            '', ''),
    ('Cisco Packet Tracer License', 'Nos', 'Cisco Packet Tracer (CSC/IT/AIML)',     '', ''),
    ('Oracle DB License',           'Nos', 'Oracle Database (MCA/IT)',              '2024-04-01', '2025-03-31'),
    ('ERP Software License',        'Nos', 'College ERP (Minerva) annual license',  '2025-04-01', '2026-03-31'),
    ('Library Management Software', 'Nos', 'LMS / KOHA annual license',             '2025-04-01', '2026-03-31'),
    ('PLC Software License',        'Nos', 'Delta Industrial Automation PLC (EIE)', '2024-04-01', '2025-03-31'),
]

# ── Vehicle items to add manually ─────────────────────────────────────────
# (item_name, unit, description, reg_no_placeholder, warranty_upto)
VEHICLE_ITEMS = [
    ('Car (Staff Transport)',           'Nos', 'Institution car for staff use',    'TN XX XXXX', ''),
    ('Van (Staff/Student Transport)',   'Nos', 'Institution van for transport',    'TN XX XXXX', ''),
    ('Auto Rickshaw',                   'Nos', 'Auto rickshaw for campus use',     'TN XX XXXX', ''),
    ('Mini Bus',                        'Nos', 'Mini bus / Tempo Traveller',       'TN XX XXXX', ''),
    ('Two Wheeler (Bike)',              'Nos', 'Motorcycle for campus use',        'TN XX XXXX', ''),
    ('Electric Vehicle / Golf Cart',   'Nos', 'Electric cart for campus mobility','TN XX XXXX', ''),
]

# Category definitions (is_asset: 1=asset 0=consumable, tracking: individual/bulk)
CATEGORY_DEFS = {
    'IT & Computer Equipment':      (1, 'individual'),
    'Audio Visual Equipment':        (1, 'individual'),
    'Electrical & Power Equipment':  (1, 'bulk'),
    'Electrical Fixtures':           (1, 'bulk'),
    'Furniture & Fixtures':          (1, 'bulk'),
    'Lab Instruments & Equipment':   (1, 'bulk'),
    'Lab Consumables & Components':  (0, 'bulk'),
    'Office Equipment':              (1, 'bulk'),
    'Safety & Facilities':           (1, 'individual'),
    'Library Resources':             (0, 'bulk'),
    'Sports & Gymnasium':            (1, 'bulk'),
    'Software Licenses':             (1, 'bulk'),
    'Vehicles':                      (1, 'individual'),
    'Miscellaneous':                 (0, 'bulk'),
}

SOFT_STORE  = 'IT Infrastructure'
VEHICLE_STORE = 'Transport Office'

# ── Unit guessing ────────────────────────────────────────────────────────────
def guess_unit(name_lower):
    if any(k in name_lower for k in ('book','ream','register','pad')):  return 'Nos'
    if any(k in name_lower for k in ('litre','ml','solution')):         return 'Litres'
    if any(k in name_lower for k in ('kg','gram','powder')):            return 'Kg'
    if any(k in name_lower for k in ('meter','cable','wire','pipe')):   return 'Meters'
    return 'Nos'

def get_category(name_lower):
    for kw, cat in CATEGORY_MAP.items():
        if kw in name_lower:
            return cat
    return 'Miscellaneous'

def normalize(s):
    """Strip extra spaces and title-case."""
    return ' '.join(str(s).strip().split()).title() if s else ''

# ── Read source data ─────────────────────────────────────────────────────────
print('Reading source file…')
wb_src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb_src['Complete Data']
src_rows = list(ws.iter_rows(values_only=True))
wb_src.close()

# Aggregate: (dept, item_normalized) → total_qty
data = defaultdict(int)        # (dept, item_norm) → qty
dept_floor = {}                # dept → floor

for r in src_rows[2:]:
    if not any(r): continue
    floor = normalize(r[1])
    dept  = normalize(r[2])
    item  = normalize(r[3])
    qty   = r[4] if isinstance(r[4], (int, float)) else 0
    if not dept or not item: continue
    data[(dept, item)] += int(qty)
    if dept not in dept_floor:
        dept_floor[dept] = floor

# Unique items (across all depts)
all_items = {}   # item_norm → category
for (dept, item), qty in data.items():
    if item not in all_items:
        all_items[item] = get_category(item.lower())

# Unique depts + add extra stores for software and vehicles
all_depts = sorted(dept_floor.keys())
extra_stores = [SOFT_STORE, VEHICLE_STORE]
all_stores = all_depts + [s for s in extra_stores if s not in all_depts]

print(f'  Departments   : {len(all_depts)}')
print(f'  Unique items  : {len(all_items)}')
print(f'  Stock rows    : {len(data)}')

# ── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# --- Styles ---
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=11)
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

SUB_FILL  = PatternFill('solid', fgColor='2E75B6')
SUB_FONT  = Font(bold=True, color='FFFFFF', size=10)

PRICE_FILL  = PatternFill('solid', fgColor='FFF2CC')   # yellow — fill me
PRICE_FONT  = Font(bold=True, color='7F6000', size=10)
EMPTY_FILL  = PatternFill('solid', fgColor='FFE7E7')   # light red — means blank

ALT_FILL = PatternFill('solid', fgColor='EBF3FB')
THIN = Side(style='thin', color='BDD7EE')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, col_widths):
    for col_idx, (hdr, width) in enumerate(col_widths, 1):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        c.fill  = HDR_FILL
        c.font  = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 30

def style_data_row(ws, row_num, values, alt=False):
    fill = ALT_FILL if alt else None
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        if fill: c.fill = fill
        c.border = BORDER
        c.alignment = Alignment(vertical='center', wrap_text=False)

# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — README
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'README'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 90

readme_lines = [
    ('STEP', 'ACTION'),
    ('1', 'Open this file and go to sheet "STOCK_FILL_PRICE"'),
    ('2', 'Fill in the "purchase_price" column (column F) for EVERY row — this is the unit cost in ₹'),
    ('  ', 'Rows highlighted in YELLOW need a price. You may use 0 only for items with no monetary value.'),
    ('3', 'For SOFTWARE rows: also fill license_valid_from and license_valid_till (columns J & K). Format: YYYY-MM-DD'),
    ('  ', 'For VEHICLE rows: fill batch_no with the vehicle Registration Number (column I). Also fill warranty_upto if known.'),
    ('4', 'Once all prices and dates are filled, save each sheet as a CSV (File → Save As → CSV):'),
    ('  ', '  a) "Categories"    → upload via Inventory → Import → Item Categories'),
    ('  ', '  b) "Supplier"      → upload via Inventory → Import → Item Suppliers'),
    ('  ', '  c) "Stores"        → upload via Inventory → Import → Item Stores'),
    ('  ', '  d) "Items"         → upload via Inventory → Import → Items'),
    ('  ', '  e) "STOCK_FILL_PRICE" (after filling) → upload via Inventory → Import → Item Stock'),
    ('5', 'Upload IN THE ORDER listed above — categories first, stock last'),
    ('  ', 'IMPORTANT: Do NOT change column headers in any sheet'),
    ('  ', 'IMPORTANT: The stock import will fail if item_name or item_category don\'t match exactly'),
    ('NOTE','purchase_price is the UNIT cost (per piece/unit). Qty × price = total value shown in MCC overview'),
    ('NOTE','Software + Vehicle rows are at the BOTTOM of STOCK_FILL_PRICE (after all physical items)'),
    ('NOTE','license_valid_till and warranty_upto drive expiry alerts in Minerva Inventory'),
    ('NOTE','All rows use supplier "MCE Stock Verification 2026" — create this first (from Supplier sheet)'),
]
ws1.cell(1,1).value = 'MCE Inventory Bulk Upload — Instructions'
ws1.cell(1,1).font  = Font(bold=True, size=14, color='1F4E79')
ws1.merge_cells('B1:B1')

for i, (step, action) in enumerate(readme_lines, 3):
    c_step   = ws1.cell(i, 1, step)
    c_action = ws1.cell(i, 2, action)
    if i == 3:
        c_step.fill   = HDR_FILL; c_step.font   = HDR_FONT
        c_action.fill = HDR_FILL; c_action.font = HDR_FONT
    elif step not in ('  ', 'NOTE'):
        c_step.font   = Font(bold=True, color='1F4E79')
        c_action.font = Font(bold=True)
    else:
        c_action.font = Font(italic=True, color='595959')
    c_action.alignment = Alignment(wrap_text=True, vertical='top')
    ws1.row_dimensions[i].height = 22

# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — Categories
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Categories')
ws2.sheet_view.showGridLines = False
cols = [('item_category',35),('description',50),('is_asset',10),('asset_tracking_mode',20)]
style_header(ws2, 1, cols)

used_cats = sorted(set(list(all_items.values()) + ['Software Licenses', 'Vehicles']))
for i, cat in enumerate(used_cats, 2):
    is_asset, mode = CATEGORY_DEFS.get(cat, (0, 'bulk'))
    style_data_row(ws2, i,
        [cat, f'{cat} — MCE physical verification 2026', is_asset, mode],
        alt=(i%2==0))

print(f'  Categories sheet: {len(used_cats)} rows')

# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — Supplier
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Supplier')
ws3.sheet_view.showGridLines = False
cols = [('item_supplier',40),('phone',15),('email',30),('address',40),
        ('contact_person_name',25),('contact_person_phone',20),
        ('contact_person_email',30),('description',50)]
style_header(ws3, 1, cols)
style_data_row(ws3, 2,
    [SUPPLIER, '', '', 'MCE Campus, Erode', 'Inventory Admin', '', '',
     'Placeholder supplier for MCE physical stock verification import'],
    alt=False)

# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — Stores (one per dept)
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Stores')
ws4.sheet_view.showGridLines = False
cols = [('item_store',45),('code',15),('description',60)]
style_header(ws4, 1, cols)

def dept_code(dept):
    words = dept.split()
    if len(words) == 1: return dept[:6].upper()
    return ''.join(w[0] for w in words if w).upper()[:8]

codes_used = {}
for i, store in enumerate(all_stores, 2):
    code = dept_code(store)
    orig = code
    n = 2
    while code in codes_used.values():
        code = orig[:6] + str(n); n += 1
    codes_used[store] = code
    floor = dept_floor.get(store, 'Campus')
    style_data_row(ws4, i,
        [store, code, f'{store} — {floor} — MCE Campus'],
        alt=(i%2==0))

print(f'  Stores sheet: {len(all_stores)} rows')

# ═══════════════════════════════════════════════════════════════════
# SHEET 5 — Items (unique items)
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Items')
ws5.sheet_view.showGridLines = False
cols = [('item_name',45),('item_category',35),('unit',12),('description',50)]
style_header(ws5, 1, cols)

sorted_items = sorted(all_items.items())
# Add software and vehicle items
extra_items = ([(name, 'Software Licenses', unit, desc) for name,unit,desc,_lf,_lt in SOFTWARE_ITEMS] +
               [(name, 'Vehicles', unit, desc) for name,unit,desc,_rn,_wu in VEHICLE_ITEMS])
for i, (item, cat) in enumerate(sorted_items, 2):
    unit = guess_unit(item.lower())
    style_data_row(ws5, i,
        [item, cat, unit, f'{item} — MCE physical verification'],
        alt=(i%2==0))
next_row = len(sorted_items) + 2
for j, (name, cat, unit, desc) in enumerate(extra_items):
    row_i = next_row + j
    c = ws5.cell(row_i, 1, name); c.border = BORDER; c.fill = PatternFill('solid', fgColor='E2EFDA')
    c = ws5.cell(row_i, 2, cat);  c.border = BORDER; c.fill = PatternFill('solid', fgColor='E2EFDA')
    c = ws5.cell(row_i, 3, unit); c.border = BORDER; c.fill = PatternFill('solid', fgColor='E2EFDA')
    c = ws5.cell(row_i, 4, desc); c.border = BORDER; c.fill = PatternFill('solid', fgColor='E2EFDA')

print(f'  Items sheet: {len(sorted_items)} physical + {len(extra_items)} software/vehicle')

# ═══════════════════════════════════════════════════════════════════
# SHEET 6 — STOCK_FILL_PRICE
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('STOCK_FILL_PRICE')
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = 'A2'

# Columns: mandatory + optional license/warranty columns
COL_HEADERS = [
    ('item_name',          30),
    ('item_category',      32),
    ('supplier_name',      32),
    ('store_name',         35),
    ('quantity',           12),
    ('purchase_price',     18),   # ← FILL THIS (mandatory)
    ('date',               14),
    ('symbol',             10),
    ('batch_no',           22),   # vehicles: registration number
    ('license_valid_from', 20),   # software: license start
    ('license_valid_till', 20),   # software: license expiry
    ('warranty_upto',      18),   # warranty end date
    ('description',        50),
]
style_header(ws6, 1, COL_HEADERS)

# Colour the price header red and extra columns teal
from openpyxl.comments import Comment
REQ_FILL   = PatternFill('solid', fgColor='C00000')   # red = mandatory fill
OPT_FILL   = PatternFill('solid', fgColor='1F7391')   # teal = fill for software/vehicle
OPT_FONT   = Font(bold=True, color='FFFFFF', size=10)

price_hdr = ws6.cell(1, 6)
price_hdr.fill = REQ_FILL
price_hdr.font = Font(bold=True, color='FFFFFF', size=11)
price_hdr.comment = Comment(
    'MANDATORY — Enter UNIT cost in Rs for each item.\n'
    'E.g. CPU=18000, Chair=2500, Ceiling Fan=1200\nUse 0 only if truly zero value.',
    author='MCE Inventory Import')

# batch_no col 9
ws6.cell(1,9).comment = Comment(
    'VEHICLES: Enter vehicle Registration Number here.\n'
    'E.g. TN33 AB 1234\nLeave blank for non-vehicle items.',
    author='MCE Inventory Import')
ws6.cell(1,9).fill = OPT_FILL; ws6.cell(1,9).font = OPT_FONT

# license_valid_from col 10
ws6.cell(1,10).fill = OPT_FILL; ws6.cell(1,10).font = OPT_FONT
ws6.cell(1,10).comment = Comment(
    'SOFTWARE: License start date. Format: YYYY-MM-DD\nLeave blank for non-software items.',
    author='MCE Inventory Import')

# license_valid_till col 11
ws6.cell(1,11).fill = OPT_FILL; ws6.cell(1,11).font = OPT_FONT
ws6.cell(1,11).comment = Comment(
    'SOFTWARE: License EXPIRY date. Format: YYYY-MM-DD\n'
    'This drives expiry alerts in Minerva Inventory.\nLeave blank for non-software items.',
    author='MCE Inventory Import')

# warranty_upto col 12
ws6.cell(1,12).fill = OPT_FILL; ws6.cell(1,12).font = OPT_FONT
ws6.cell(1,12).comment = Comment(
    'Warranty expiry date. Format: YYYY-MM-DD\n'
    'Applicable for equipment, vehicles, instruments.\nLeave blank if not known.',
    author='MCE Inventory Import')

SOFT_GREEN  = PatternFill('solid', fgColor='E2EFDA')   # light green for software rows
VEH_ORANGE  = PatternFill('solid', fgColor='FCE4D6')   # light orange for vehicle rows

row_num = 2
sorted_stock = sorted(data.items(), key=lambda x: (x[0][0], x[0][1]))

def write_stock_row(ws, row_num, item, cat, store, qty, date_val,
                   batch_no='', lic_from='', lic_till='', warranty='',
                   row_fill=None):
    vals = [item, cat, SUPPLIER, store, qty, None, date_val, '+',
            batch_no, lic_from, lic_till, warranty,
            f'{item} — {store} — MCE stock verification 2026']
    for col, val in enumerate(vals, 1):
        c = ws.cell(row_num, col, val)
        c.border = BORDER
        c.alignment = Alignment(vertical='center')
        if row_fill:
            c.fill = row_fill
        elif row_num % 2 == 0:
            c.fill = ALT_FILL
    # Price cell always yellow
    pc = ws.cell(row_num, 6)
    pc.fill = PRICE_FILL; pc.font = PRICE_FONT; pc.value = None

for (dept, item), qty in sorted_stock:
    write_stock_row(ws6, row_num, item, all_items.get(item, 'Miscellaneous'), dept, qty, IMPORT_DATE)
    row_num += 1

# ── Software rows (green) ─────────────────────────────────────────
phys_end = row_num
for name, unit, desc, lic_from, lic_till in SOFTWARE_ITEMS:
    write_stock_row(ws6, row_num, name, 'Software Licenses', SOFT_STORE,
                   1, IMPORT_DATE, '', lic_from, lic_till, '', SOFT_GREEN)
    row_num += 1

# ── Vehicle rows (orange) ─────────────────────────────────────────
for name, unit, desc, reg_placeholder, warranty in VEHICLE_ITEMS:
    write_stock_row(ws6, row_num, name, 'Vehicles', VEHICLE_STORE,
                   1, IMPORT_DATE, reg_placeholder, '', '', warranty, VEH_ORANGE)
    row_num += 1

# Add auto-filter
ws6.auto_filter.ref = f'A1:{get_column_letter(len(COL_HEADERS))}1'

soft_count = len(SOFTWARE_ITEMS)
veh_count  = len(VEHICLE_ITEMS)
phys_count = phys_end - 2
print(f'  Stock sheet: {phys_count} physical + {soft_count} software + {veh_count} vehicle rows')

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f'\n✓ Saved: {OUT}')
print('\nSummary:')
print(f'  Stores               : {len(all_stores)} ({len(all_depts)} depts + 2 extra)')
print(f'  Categories           : {len(used_cats)}')
print(f'  Items (physical)     : {len(sorted_items)}')
print(f'  Items (software+veh) : {len(extra_items)}')
print(f'  Stock rows           : {phys_count} physical + {soft_count} software + {veh_count} vehicle')
print(f'  Supplier             : {SUPPLIER}')
print(f'  Import date          : {IMPORT_DATE}')
