from collections import Counter
from pathlib import Path
import json

from openpyxl import load_workbook


WORKSPACE = Path("/Applications/XAMPP/xamppfiles/htdocs/minerva")
WORKBOOK_GLOB = "Review comments - Jan*2026*New*.xlsx"
OUT_JSON = WORKSPACE / "docs" / "review_comments_jan2026_analysis.json"

COLS = {
    "month": 1,
    "vertical": 2,
    "property_code": 3,
    "property_name": 4,
    "review_comments": 5,
    "category": 7,
    "sub_category": 8,
    "accountant": 9,
    "reviewer": 10,
    "ac_name": 11,
    "controller": 12,
    "criticality": 13,
    "manager": 14,
}


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def to_common_list(counter_obj, limit=15):
    return [{"label": k, "count": v} for k, v in counter_obj.most_common(limit)]


def analyze_sheet(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {k: clean(ws.cell(r, c).value) for k, c in COLS.items()}
        if not any([row["property_code"], row["property_name"], row["review_comments"], row["category"]]):
            continue
        rows.append(row)

    total = len(rows)
    critical_rows = [x for x in rows if x["criticality"].lower() == "critical"]
    non_critical_rows = [x for x in rows if x["criticality"].lower() == "non critical"]

    return {
        "rows": rows,
        "summary": {
            "total_rows": total,
            "critical_count": len(critical_rows),
            "non_critical_count": len(non_critical_rows),
            "critical_pct": round((len(critical_rows) / total) * 100, 2) if total else 0,
            "top_categories": to_common_list(Counter(x["category"] or "Unspecified" for x in rows), 12),
            "top_sub_categories": to_common_list(Counter(x["sub_category"] or "Unspecified" for x in rows), 12),
            "top_controllers": to_common_list(Counter(x["controller"] or "Unspecified" for x in rows), 10),
            "top_reviewers": to_common_list(Counter(x["reviewer"] or "Unspecified" for x in rows), 10),
            "top_managers": to_common_list(Counter(x["manager"] or "Unspecified" for x in rows), 10),
            "top_properties_by_volume": [
                {"property_code": k[0], "property_name": k[1], "count": v}
                for k, v in Counter((x["property_code"], x["property_name"]) for x in rows).most_common(12)
            ],
            "critical_hotspots": [
                {"property_code": k[0], "property_name": k[1], "count": v}
                for k, v in Counter((x["property_code"], x["property_name"]) for x in critical_rows).most_common(12)
            ],
        },
    }


def main():
    workbook = next(WORKSPACE.glob(WORKBOOK_GLOB), None)
    if workbook is None:
        raise FileNotFoundError(f"Workbook not found using pattern: {WORKBOOK_GLOB}")

    wb = load_workbook(workbook, data_only=True)

    per_sheet = {}
    all_rows = []

    for sheet_name in wb.sheetnames:
        result = analyze_sheet(wb[sheet_name])
        per_sheet[sheet_name] = result["summary"]
        all_rows.extend(result["rows"])

    combined_total = len(all_rows)
    combined_critical = [x for x in all_rows if x["criticality"].lower() == "critical"]
    combined_non_critical = [x for x in all_rows if x["criticality"].lower() == "non critical"]

    combined = {
        "total_rows": combined_total,
        "critical_count": len(combined_critical),
        "non_critical_count": len(combined_non_critical),
        "critical_pct": round((len(combined_critical) / combined_total) * 100, 2) if combined_total else 0,
        "sheet_breakup": {k: v["total_rows"] for k, v in per_sheet.items()},
        "top_categories": to_common_list(Counter(x["category"] or "Unspecified" for x in all_rows), 15),
        "top_sub_categories": to_common_list(Counter(x["sub_category"] or "Unspecified" for x in all_rows), 15),
        "top_controllers": to_common_list(Counter(x["controller"] or "Unspecified" for x in all_rows), 12),
        "top_reviewers": to_common_list(Counter(x["reviewer"] or "Unspecified" for x in all_rows), 12),
        "top_managers": to_common_list(Counter(x["manager"] or "Unspecified" for x in all_rows), 12),
        "top_properties_by_volume": [
            {"property_code": k[0], "property_name": k[1], "count": v}
            for k, v in Counter((x["property_code"], x["property_name"]) for x in all_rows).most_common(15)
        ],
        "critical_hotspots": [
            {"property_code": k[0], "property_name": k[1], "count": v}
            for k, v in Counter((x["property_code"], x["property_name"]) for x in combined_critical).most_common(15)
        ],
    }

    payload = {
        "workbook": str(workbook),
        "sheets": wb.sheetnames,
        "sheet_results": per_sheet,
        "combined": combined,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(str(OUT_JSON))


if __name__ == "__main__":
    main()
