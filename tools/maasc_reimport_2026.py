#!/usr/bin/env python3
"""
maasc_reimport_2026.py  —  Delete all existing maasc students and re-import for 2026-27.

Steps:
  1. DELETE all existing students, student_sessions, student/parent users
  2. INSERT new classes, sections, class_sections as needed
  3. INSERT students from maasc_student_details/ (15 Excel files)
  4. INSERT student_session records (session_id=22 / 2026-27)
  5. INSERT student and parent user accounts

Run:
  cd /Applications/XAMPP/xamppfiles/htdocs/minerva
  source .venv/bin/activate
  python3 tools/maasc_reimport_2026.py           # just generate SQL
  python3 tools/maasc_reimport_2026.py --deploy  # generate + SCP + execute on EC2
"""

import os
import re
import shutil
import subprocess
import sys
import calendar
from datetime import datetime, date

import openpyxl

BASE     = '/Applications/XAMPP/xamppfiles/htdocs/minerva'
DATA_DIR = f'{BASE}/maasc_student_details'
OUT_SQL  = f'{BASE}/tools/maasc_reimport_2026.sql'

# ── DB constants ──────────────────────────────────────────────────────────────
SESSION_ID   = 22       # 2026-27
LANG_ID      = 4        # English
CURRENCY_ID  = 68       # INR
BATCH_TAG    = 'IMP_MAASC_2026'
BCRYPT_PW    = r'$2y$12$mUoUzS1Fgc0a.qsXU1jIf.V37EhCavy82beMW9K9Zd9lqFQ9NMgyK'  # Welcome@123

# ── File configs ──────────────────────────────────────────────────────────────
# fmt: 'standard' = lowercase headers directly usable
#      'bcom'     = Title-Case limited headers (no guardian/middlename)
#      'mcom'     = special verbose headers, header at row 3, data at row 4
FILE_CONFIGS = [
    {'file': 'II B.COM A&F.xlsx',       'class': 'II B.COM',          'section': 'A&F', 'header_row': 1, 'data_row': 2, 'fmt': 'bcom'},
    {'file': 'II B.COM GENERAL.xlsx',   'class': 'II B.COM',          'section': 'B',   'header_row': 1, 'data_row': 2, 'fmt': 'bcom'},
    {'file': 'II B.Sc CHEM.xlsx',       'class': 'II B.Sc Chemistry', 'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'II BA TAMIL.XLS',         'class': 'II BA Tamil',       'section': 'A',   'header_row': 1, 'data_row': 3, 'fmt': 'standard'},
    {'file': 'II BBA.xlsx',             'class': 'II BBA',            'section': 'BBA', 'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'II BCA.xlsx',             'class': 'II BCA',            'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'II CS.xlsx',              'class': 'II B.SC CS',        'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'II M.COM GENERAL.xlsx',   'class': 'II M.Com',          'section': 'A',   'header_row': 3, 'data_row': 4, 'fmt': 'mcom'},
    {'file': 'III B.Sc CHEM.xlsx',      'class': 'III CHEMISTRY',     'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'III B.Sc MATHS.xlsx',     'class': 'III MATHS',         'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'III B.Sc PHYSICS.xlsx',   'class': 'III B.Sc PHYSICS',  'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'III BA TAMIL.XLS',        'class': 'III BA Tamil',      'section': 'A',   'header_row': 1, 'data_row': 3, 'fmt': 'standard'},
    {'file': 'III BBA.xlsx',            'class': 'III BBA',           'section': 'BBA', 'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'III BCA.xlsx',            'class': 'III BCA',           'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
    {'file': 'III CS.xlsx',             'class': 'III B.SC CS',       'section': 'A',   'header_row': 1, 'data_row': 2, 'fmt': 'standard'},
]

# MCOM column map: normalized_header → standard_field
MCOM_COL_MAP = {
    'application number':    'admission_no',
    'register number':       'register_no',
    'name':                  '_fullname',
    'gender':                'gender',
    'date of birth':         'dob',
    'religion':              'religion',
    'community':             'cast',
    'community category i':  '_cast_cat',
    'student mobile numbe':  'mobileno',   # header is truncated in file
    'student mobile number': 'mobileno',
    'e- mail id':            'email',
    'e-mail id':             'email',
    'admission date':        'admission_date',
    'blood group':           'blood_group',
    'height':                'height',
    'weight':                'weight',
    'measurement date':      'measurement_date',
    'father name':           'father_name',
    'father phone  number':  'father_phone',
    'father phone number':   'father_phone',
    'father occupation':     'father_occupation',
    'father aadhar number':  'father_adhar_no',
    'm0ther name':           'mother_name',    # typo in source (zero not O)
    'mother name':           'mother_name',
    'mother phone  number':  'mother_phone',
    'mother phone number':   'mother_phone',
    'mother occupation':     'mother_occupation',
    'mother aadhar number':  'mother_adhar_no',
    'guardian':              'guardian_is',
    'guardian name':         'guardian_name',
    'guardian phone':        'guardian_phone',
    'guardian occupation':   'guardian_occupation',
    'guardian address':      'guardian_address',
    'current address':       'current_address',
    'permanent address':     'permanent_address',
    'bank name':             'bank_name',
    'ifsc code':             'ifsc_code',
    'aadhar number':         'adhar_no',
    'previous school name':  'previous_school',
    'rte id  ( right to e':  'rte',
    'rte id':                'rte',
    'regulation id':         'regulation_id',
    'emis number':           'emis_num',
    'higher secondary reg':  'hsc_reg_no',
    'ug register number':    'ug_reg_no',
    'abc id (academic ban':  'abc_id',
    'abc id':                'abc_id',
    ' migration certifica':  'migration_cert_num',
    'migration certificate': 'migration_cert_num',
}

# ── SQL helpers ───────────────────────────────────────────────────────────────
lines = []

def sql(s: str):
    lines.append(s)

def esc(v) -> str:
    if v is None:
        return "''"
    s = str(v).strip()
    s = s.replace('\\', '\\\\').replace("'", "\\'")
    return f"'{s}'"

def esc_null(v) -> str:
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if s.lower() in ('', 'none', 'nan', 'null', 'na', 'n/a', '-', '0', 'late'):
        return 'NULL'
    s = s.replace('\\', '\\\\').replace("'", "\\'")
    return f"'{s}'"

def clean_str(v) -> str:
    """Normalise a cell value to string, converting float ints (474.0 → '474')."""
    if v is None:
        return ''
    s = str(v).strip()
    if s.lower() in ('none', 'nan', 'null', 'na', 'n/a', '-'):
        return ''
    # Excel stores plain integers as floats (e.g. 474.0)
    if re.match(r'^\d+\.0+$', s):
        return str(int(float(s)))
    return s

def clean_null(v):
    s = clean_str(v)
    return s if s else None

def clean_aadhaar12(v) -> str | None:
    """Strip spaces/hyphens from Aadhaar and keep at most 12 chars (varchar(12) limit)."""
    s = clean_str(v)
    if not s:
        return None
    digits = re.sub(r'[\s\-]', '', s)   # remove spaces and hyphens
    return digits[:12] if digits else None

def clean_mobile(v) -> str | None:
    """Convert a phone number cell to string, stripping .0 suffix."""
    s = clean_str(v)
    if not s:
        return None
    # Already clean digits
    if re.match(r'^\d{7,15}$', s):
        return s
    # Float like 9789355487.0
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s or None

def _valid_date(y: int, mo: int, d: int) -> bool:
    """Return True only if year/month/day form a real calendar date."""
    if not (1900 <= y <= 2100 and 1 <= mo <= 12):
        return False
    return 1 <= d <= calendar.monthrange(y, mo)[1]

def fmt_date(v) -> str:
    """Return SQL date literal 'YYYY-MM-DD' or NULL for blank/invalid/datetime."""
    if v is None:
        return 'NULL'
    # openpyxl datetime object
    if isinstance(v, (datetime, date)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).strip()
    # Strip time component and embedded DOB annotations
    s = s.split(' ')[0].split('\n')[0].split('[')[0].strip()
    if not s or s.lower() in ('none', 'nan', 'null', 'na', 'n/a', '-', '', '0'):
        return 'NULL'

    # YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
    if m:
        y, mo, d = int(m[1]), int(m[2]), int(m[3])
        return f"'{y:04d}-{mo:02d}-{d:02d}'" if _valid_date(y, mo, d) else 'NULL'

    # DD.MM.YYYY or D.M.YYYY
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', s)
    if m:
        d, mo, y = int(m[1]), int(m[2]), int(m[3])
        return f"'{y:04d}-{mo:02d}-{d:02d}'" if _valid_date(y, mo, d) else 'NULL'

    # DD/MM/YYYY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = int(m[1]), int(m[2]), int(m[3])
        return f"'{y:04d}-{mo:02d}-{d:02d}'" if _valid_date(y, mo, d) else 'NULL'

    # DD-MM-YYYY
    m = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', s)
    if m:
        d, mo, y = int(m[1]), int(m[2]), int(m[3])
        return f"'{y:04d}-{mo:02d}-{d:02d}'" if _valid_date(y, mo, d) else 'NULL'

    # YYYY.MM.DD  (or YYYY.DD.MM when month > 12)
    m = re.match(r'^(\d{4})\.(\d{1,2})\.(\d{1,2})$', s)
    if m:
        y, p2, p3 = int(m[1]), int(m[2]), int(m[3])
        if _valid_date(y, p2, p3):
            return f"'{y:04d}-{p2:02d}-{p3:02d}'"
        if _valid_date(y, p3, p2):   # swap (e.g. 2024.13.06 → 2024-06-13)
            return f"'{y:04d}-{p3:02d}-{p2:02d}'"
        return 'NULL'

    # YYYY/MM/DD
    m = re.match(r'^(\d{4})/(\d{2})/(\d{2})$', s)
    if m:
        y, mo, d = int(m[1]), int(m[2]), int(m[3])
        return f"'{y:04d}-{mo:02d}-{d:02d}'" if _valid_date(y, mo, d) else 'NULL'

    return 'NULL'

def norm_header(h: str) -> str:
    """Normalise a column header: lowercase, strip, fix known quirks."""
    if not h:
        return ''
    h = h.lower().strip()
    h = re.sub(r'_\s+', '_', h)    # 'admission_ no' → 'admission_no'
    h = re.sub(r'\s+_', '_', h)    # trailing-space before underscore
    return h

# ── File loaders ──────────────────────────────────────────────────────────────

def open_wb(filepath: str):
    """Open workbook, handling .XLS (xlsx-disguised) by copying to temp .xlsx."""
    if filepath.upper().endswith('.XLS'):
        tmp = filepath + '_tmp_read.xlsx'
        shutil.copy(filepath, tmp)
        try:
            wb = openpyxl.load_workbook(tmp, data_only=True)
        finally:
            os.remove(tmp)
        return wb
    return openpyxl.load_workbook(filepath, data_only=True)

def ws_rows(ws, header_row: int, data_row: int):
    """Yield row dicts from worksheet."""
    # Read headers
    raw_headers = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        raw_headers = [norm_header(str(c.value) if c.value else '') for c in row]
        break
    # Yield data rows as dicts
    for row in ws.iter_rows(min_row=data_row):
        vals = [c.value for c in row]
        if not any(v for v in vals):
            continue   # skip empty rows
        d = {}
        for i, h in enumerate(raw_headers):
            if h and i < len(vals):
                d[h] = vals[i]
        yield d

def load_standard(filepath, class_name, section_name, header_row, data_row):
    """Load a standard-format file (lowercase headers)."""
    records = []
    wb = open_wb(filepath)
    ws = wb.active
    for d in ws_rows(ws, header_row, data_row):
        # Normalise 'caste' → 'cast'
        if 'caste' in d and 'cast' not in d:
            d['cast'] = d.pop('caste')
        # admission_no: might be in 'admission_no' or 'admission__no' after normalisation
        adm = clean_str(d.get('admission_no') or d.get('admission__no') or '')
        firstname = clean_str(d.get('firstname', ''))
        if not adm or not firstname:
            continue
        records.append({
            'admission_no':      adm,
            'roll_no':           clean_null(d.get('roll_no')),
            'class_name':        class_name,
            'section_name':      section_name,
            'firstname':         firstname,
            'middlename':        clean_null(d.get('middlename')),
            'lastname':          clean_null(d.get('lastname')),
            'gender':            clean_str(d.get('gender', '')),
            'dob':               fmt_date(d.get('dob')),
            'category_id':       clean_null(d.get('category_id')),
            'religion':          clean_null(d.get('religion')),
            'cast':              clean_null(d.get('cast')),
            'mobileno':          clean_mobile(d.get('mobileno')),
            'email':             clean_null(d.get('email')),
            'admission_date':    fmt_date(d.get('admission_date')),
            'blood_group':       clean_str(d.get('blood_group', '')),
            'height':            clean_str(d.get('height', '')),
            'weight':            clean_str(d.get('weight', '')),
            'measurement_date':  fmt_date(d.get('measurement_date')),
            'father_name':       clean_null(d.get('father_name')),
            'father_phone':      clean_mobile(d.get('father_phone')),
            'father_occupation': clean_null(d.get('father_occupation')),
            'mother_name':       clean_null(d.get('mother_name')),
            'mother_phone':      clean_mobile(d.get('mother_phone')),
            'mother_occupation': clean_null(d.get('mother_occupation')),
            'guardian_is':       clean_str(d.get('guardian_is', 'father')) or 'father',
            'guardian_name':     clean_null(d.get('guardian_name')),
            'guardian_relation': clean_null(d.get('guardian_relation')),
            'guardian_email':    clean_null(d.get('guardian_email')),
            'guardian_phone':    clean_mobile(d.get('guardian_phone')),
            'guardian_occupation': clean_null(d.get('guardian_occupation')),
            'guardian_address':  clean_null(d.get('guardian_address')),
            'current_address':   clean_null(d.get('current_address')),
            'permanent_address': clean_null(d.get('permanent_address')),
            'bank_account_no':   clean_null(d.get('bank_account_no')),
            'bank_name':         clean_null(d.get('bank_name')),
            'ifsc_code':         clean_null(d.get('ifsc_code')),
            'adhar_no':          clean_null(d.get('adhar_no')),
            'father_adhar_no':   clean_aadhaar12(d.get('father_adhar_no')),
            'mother_adhar_no':   clean_aadhaar12(d.get('mother_adhar_no')),
            'register_no':       clean_null(d.get('register_no')),
            'regulation_id':     clean_null(d.get('regulation_id')),
            'emis_num':          clean_null(d.get('emis_num')),
            'hsc_reg_no':        clean_null(d.get('hsc_reg_no')),
            'ug_reg_no':         clean_null(d.get('ug_reg_no')),
            'abc_id':            clean_null(d.get('abc_id')),
            'migration_cert_num': clean_null(d.get('migration_cert_num')),
            'rte':               clean_null(d.get('rte')),
            'medium':            clean_null(d.get('medium')),
            'samagra_id':        clean_null(d.get('samagra_id')),
            'previous_school':   clean_null(d.get('previous_school')),
        })
    wb.close()
    return records

def load_bcom(filepath, class_name, section_name, header_row, data_row):
    """Load a B.COM-style file (Title-Case headers, limited fields)."""
    records = []
    wb = open_wb(filepath)
    ws = wb.active
    for d in ws_rows(ws, header_row, data_row):
        # Title-Case headers after norm_header are lowercased; 'caste' → 'cast'
        if 'caste' in d and 'cast' not in d:
            d['cast'] = d.pop('caste')
        adm = clean_str(d.get('admission_no', ''))
        firstname = clean_str(d.get('firstname', ''))
        if not adm or not firstname:
            continue
        records.append({
            'admission_no':      adm,
            'roll_no':           clean_null(d.get('roll_no')),
            'class_name':        class_name,
            'section_name':      section_name,
            'firstname':         firstname,
            'middlename':        None,
            'lastname':          clean_null(d.get('lastname')),
            'gender':            clean_str(d.get('gender', '')),
            'dob':               fmt_date(d.get('dob')),
            'category_id':       None,
            'religion':          clean_null(d.get('religion')),
            'cast':              clean_null(d.get('cast')),
            'mobileno':          clean_mobile(d.get('mobileno')),
            'email':             clean_null(d.get('email')),
            'admission_date':    fmt_date(d.get('admission_date')),
            'blood_group':       clean_str(d.get('blood_group', '')),
            'height':            clean_str(d.get('height', '')),
            'weight':            clean_str(d.get('weight', '')),
            'measurement_date':  fmt_date(d.get('measurement_date')),
            'father_name':       clean_null(d.get('father_name')),
            'father_phone':      clean_mobile(d.get('father_phone')),
            'father_occupation': clean_null(d.get('father_occupation')),
            'mother_name':       clean_null(d.get('mother_name')),
            'mother_phone':      clean_mobile(d.get('mother_phone')),
            'mother_occupation': clean_null(d.get('mother_occupation')),
            'guardian_is':       'father',
            'guardian_name':     None,
            'guardian_relation': None,
            'guardian_email':    None,
            'guardian_phone':    None,
            'guardian_occupation': None,
            'guardian_address':  None,
            'current_address':   clean_null(d.get('current_address')),
            'permanent_address': clean_null(d.get('permanent_address')),
            'bank_account_no':   clean_null(d.get('bank_account_no')),
            'bank_name':         clean_null(d.get('bank_name')),
            'ifsc_code':         clean_null(d.get('ifsc_code')),
            'adhar_no':          clean_null(d.get('adhar_no')),
            'father_adhar_no':   clean_aadhaar12(d.get('father_adhar_no')),
            'mother_adhar_no':   clean_aadhaar12(d.get('mother_adhar_no')),
            'register_no':       None,
            'regulation_id':     None,
            'emis_num':          clean_null(d.get('emis_num')),
            'hsc_reg_no':        clean_null(d.get('hsc_reg_no')),
            'ug_reg_no':         clean_null(d.get('ug_reg_no')),
            'abc_id':            None,
            'migration_cert_num': None,
            'rte':               None,
            'medium':            clean_null(d.get('medium')),
            'samagra_id':        None,
            'previous_school':   clean_null(d.get('previous_school')),
        })
    wb.close()
    return records

def load_mcom(filepath, class_name, section_name, header_row, data_row):
    """Load M.COM GENERAL file (verbose headers, header at row 3, data at row 4)."""
    records = []
    wb = open_wb(filepath)
    ws = wb.active
    for d in ws_rows(ws, header_row, data_row):
        # Map verbose header keys to standard field names
        mapped = {}
        for raw_key, val in d.items():
            std_field = MCOM_COL_MAP.get(raw_key.strip())
            if std_field and std_field not in ('_cast_cat',):
                mapped[std_field] = val
            elif std_field == '_cast_cat' and 'cast' not in mapped:
                mapped['cast'] = val  # fallback: use community category I as cast
        # Parse full name (may contain "\n[DOB]")
        full_name = clean_str(mapped.get('_fullname', ''))
        if not full_name:
            full_name = clean_str(mapped.get('admission_no', ''))  # won't be right, just guard
        full_name = re.sub(r'\s*[\n\[].+', '', full_name).strip()
        adm = clean_str(mapped.get('admission_no', ''))
        if not adm or not full_name:
            continue
        records.append({
            'admission_no':      adm,
            'roll_no':           None,
            'class_name':        class_name,
            'section_name':      section_name,
            'firstname':         full_name,
            'middlename':        None,
            'lastname':          None,
            'gender':            clean_str(mapped.get('gender', '')),
            'dob':               fmt_date(mapped.get('dob')),
            'category_id':       None,
            'religion':          clean_null(mapped.get('religion')),
            'cast':              clean_null(mapped.get('cast')),
            'mobileno':          clean_mobile(mapped.get('mobileno')),
            'email':             clean_null(mapped.get('email')),
            'admission_date':    fmt_date(mapped.get('admission_date')),
            'blood_group':       clean_str(mapped.get('blood_group', '')),
            'height':            clean_str(mapped.get('height', '')),
            'weight':            clean_str(mapped.get('weight', '')),
            'measurement_date':  fmt_date(mapped.get('measurement_date')),
            'father_name':       clean_null(mapped.get('father_name')),
            'father_phone':      clean_mobile(mapped.get('father_phone')),
            'father_occupation': clean_null(mapped.get('father_occupation')),
            'mother_name':       clean_null(mapped.get('mother_name')),
            'mother_phone':      clean_mobile(mapped.get('mother_phone')),
            'mother_occupation': clean_null(mapped.get('mother_occupation')),
            'guardian_is':       clean_str(mapped.get('guardian_is', 'father')) or 'father',
            'guardian_name':     clean_null(mapped.get('guardian_name')),
            'guardian_relation': None,
            'guardian_email':    None,
            'guardian_phone':    clean_mobile(mapped.get('guardian_phone')),
            'guardian_occupation': clean_null(mapped.get('guardian_occupation')),
            'guardian_address':  clean_null(mapped.get('guardian_address')),
            'current_address':   clean_null(mapped.get('current_address')),
            'permanent_address': clean_null(mapped.get('permanent_address')),
            'bank_account_no':   None,
            'bank_name':         clean_null(mapped.get('bank_name')),
            'ifsc_code':         clean_null(mapped.get('ifsc_code')),
            'adhar_no':          clean_null(mapped.get('adhar_no')),
            'father_adhar_no':   clean_aadhaar12(mapped.get('father_adhar_no')),
            'mother_adhar_no':   clean_aadhaar12(mapped.get('mother_adhar_no')),
            'register_no':       clean_null(mapped.get('register_no')),
            'regulation_id':     clean_null(mapped.get('regulation_id')),
            'emis_num':          clean_null(mapped.get('emis_num')),
            'hsc_reg_no':        clean_null(mapped.get('hsc_reg_no')),
            'ug_reg_no':         clean_null(mapped.get('ug_reg_no')),
            'abc_id':            clean_null(mapped.get('abc_id')),
            'migration_cert_num': clean_null(mapped.get('migration_cert_num')),
            'rte':               clean_null(mapped.get('rte')),
            'medium':            None,
            'samagra_id':        None,
            'previous_school':   clean_null(mapped.get('previous_school')),
        })
    wb.close()
    return records

# ─────────────────────────────────────────────────────────────────────────────
# LOAD ALL FILES
# ─────────────────────────────────────────────────────────────────────────────
all_records = []
for cfg in FILE_CONFIGS:
    fpath = os.path.join(DATA_DIR, cfg['file'])
    if not os.path.exists(fpath):
        print(f"WARNING: file not found: {fpath}", file=sys.stderr)
        continue
    fmt = cfg['fmt']
    if fmt == 'standard':
        recs = load_standard(fpath, cfg['class'], cfg['section'], cfg['header_row'], cfg['data_row'])
    elif fmt == 'bcom':
        recs = load_bcom(fpath, cfg['class'], cfg['section'], cfg['header_row'], cfg['data_row'])
    elif fmt == 'mcom':
        recs = load_mcom(fpath, cfg['class'], cfg['section'], cfg['header_row'], cfg['data_row'])
    else:
        recs = []
    print(f"  {cfg['file']}: {len(recs)} students loaded")
    all_records.extend(recs)

print(f"\nTotal students loaded: {len(all_records)}")

# Unique classes and sections
class_section_pairs = list(dict.fromkeys(
    (r['class_name'], r['section_name']) for r in all_records
))
unique_classes   = sorted(set(p[0] for p in class_section_pairs))
unique_sections  = sorted(set(p[1] for p in class_section_pairs))
print(f"Classes: {unique_classes}")
print(f"Sections: {unique_sections}")

# ─────────────────────────────────────────────────────────────────────────────
# GENERATE SQL
# ─────────────────────────────────────────────────────────────────────────────

sql("-- ============================================================")
sql("-- maasc_reimport_2026.sql  —  generated by tools/maasc_reimport_2026.py")
sql("-- Institution: MAASC   DB: maasc   Session: 2026-27 (id=22)")
sql(f"-- Batch tag: {BATCH_TAG}   Students: {len(all_records)}")
sql("-- ============================================================")
sql("SET NAMES utf8mb4;")
sql("SET FOREIGN_KEY_CHECKS = 0;")
sql("")

# ─── 1. DELETE EXISTING STUDENTS ─────────────────────────────────────────────
sql("-- ── 1. Delete all existing student/parent users and student data ──────")
sql("DELETE FROM users WHERE role IN ('student', 'parent');")
sql("DELETE FROM student_session WHERE 1=1;")
sql("DELETE FROM students WHERE 1=1;")
sql("")

# ─── 2. CLASSES ───────────────────────────────────────────────────────────────
sql("-- ── 2. Classes (create new ones, existing ones are preserved) ─────────")
for cls in unique_classes:
    sql(f"INSERT INTO classes (class, is_active, class_type)"
        f" SELECT {esc(cls)}, 'yes', 'academic'"
        f" WHERE NOT EXISTS (SELECT 1 FROM classes WHERE LOWER(class)=LOWER({esc(cls)}));")
sql("")

# ─── 3. SECTIONS ─────────────────────────────────────────────────────────────
sql("-- ── 3. Sections ───────────────────────────────────────────────────────")
for sec in unique_sections:
    sql(f"INSERT INTO sections (section, is_active)"
        f" SELECT {esc(sec)}, 'yes'"
        f" WHERE NOT EXISTS (SELECT 1 FROM sections WHERE LOWER(section)=LOWER({esc(sec)}));")
sql("")

# ─── 4. CLASS_SECTIONS ───────────────────────────────────────────────────────
sql("-- ── 4. Class-section links ────────────────────────────────────────────")
for cls, sec in class_section_pairs:
    sql(f"INSERT INTO class_sections (class_id, section_id, is_active)"
        f" SELECT c.id, s.id, 'yes'"
        f" FROM classes c, sections s"
        f" WHERE LOWER(c.class)=LOWER({esc(cls)}) AND LOWER(s.section)=LOWER({esc(sec)})"
        f" AND NOT EXISTS ("
        f"   SELECT 1 FROM class_sections cs2"
        f"   JOIN classes cc ON cc.id=cs2.class_id"
        f"   JOIN sections ss ON ss.id=cs2.section_id"
        f"   WHERE LOWER(cc.class)=LOWER({esc(cls)}) AND LOWER(ss.section)=LOWER({esc(sec)})"
        f");")
sql("")

# ─── 5. STUDENTS ─────────────────────────────────────────────────────────────
sql("-- ── 5. Students ──────────────────────────────────────────────────────")

for r in all_records:
    adm_no = r['admission_no']
    sql(f"-- {r['class_name']}: {r['firstname']}")
    sql(
        f"INSERT INTO students ("
        f"admission_no, roll_no, firstname, middlename, lastname, "
        f"gender, dob, category_id, religion, cast, "
        f"mobileno, email, admission_date, blood_group, "
        f"height, weight, measurement_date, "
        f"father_name, father_phone, father_occupation, "
        f"mother_name, mother_phone, mother_occupation, "
        f"guardian_is, guardian_name, guardian_relation, guardian_email, "
        f"guardian_phone, guardian_occupation, guardian_address, "
        f"current_address, permanent_address, "
        f"bank_account_no, bank_name, ifsc_code, "
        f"adhar_no, father_adhar_no, mother_adhar_no, "
        f"register_no, regulation_id, emis_num, hsc_reg_no, ug_reg_no, "
        f"abc_id, migration_cert_num, rte, medium, samagra_id, "
        f"previous_school, "
        f"is_active, parent_id, dis_note, father_pic, mother_pic, guardian_pic, note"
        f")"
        f" SELECT "
        f"{esc(adm_no)}, {esc_null(r['roll_no'])}, {esc(r['firstname'])}, "
        f"{esc_null(r['middlename'])}, {esc_null(r['lastname'])}, "
        f"{esc(r['gender'])}, {r['dob']}, {esc_null(r['category_id'])}, "
        f"{esc_null(r['religion'])}, {esc_null(r['cast'])}, "
        f"{esc_null(r['mobileno'])}, {esc_null(r['email'])}, "
        f"{r['admission_date']}, {esc(r['blood_group'])}, "
        f"{esc(r['height'])}, {esc(r['weight'])}, {r['measurement_date']}, "
        f"{esc_null(r['father_name'])}, {esc_null(r['father_phone'])}, {esc_null(r['father_occupation'])}, "
        f"{esc_null(r['mother_name'])}, {esc_null(r['mother_phone'])}, {esc_null(r['mother_occupation'])}, "
        f"{esc(r['guardian_is'])}, {esc_null(r['guardian_name'])}, {esc_null(r['guardian_relation'])}, "
        f"{esc_null(r['guardian_email'])}, {esc_null(r['guardian_phone'])}, "
        f"{esc(r['guardian_occupation'] or '')}, {esc_null(r['guardian_address'])}, "
        f"{esc_null(r['current_address'])}, {esc_null(r['permanent_address'])}, "
        f"{esc_null(r['bank_account_no'])}, {esc_null(r['bank_name'])}, {esc_null(r['ifsc_code'])}, "
        f"{esc_null(r['adhar_no'])}, {esc_null(r['father_adhar_no'])}, {esc_null(r['mother_adhar_no'])}, "
        f"{esc_null(r['register_no'])}, {esc_null(r['regulation_id'])}, {esc_null(r['emis_num'])}, "
        f"{esc_null(r['hsc_reg_no'])}, {esc_null(r['ug_reg_no'])}, "
        f"{esc_null(r['abc_id'])}, {esc_null(r['migration_cert_num'])}, "
        f"{esc_null(r['rte'])}, {esc_null(r['medium'])}, {esc_null(r['samagra_id'])}, "
        f"{esc_null(r['previous_school'])}, "
        f"'yes', 0, '', '', '', '', {esc(BATCH_TAG)} "
        f"WHERE NOT EXISTS (SELECT 1 FROM students WHERE admission_no={esc(adm_no)});"
    )

sql("")

# ─── 6. STUDENT SESSIONS ─────────────────────────────────────────────────────
sql("-- ── 6. Student sessions ──────────────────────────────────────────────")
for cls, sec in class_section_pairs:
    cls_adm_nos = [r['admission_no'] for r in all_records
                   if r['class_name'] == cls and r['section_name'] == sec]
    if not cls_adm_nos:
        continue
    sql(f"-- {cls} / {sec}")
    sql(f"INSERT INTO student_session (session_id, student_id, class_id, section_id, is_active, is_alumni)")
    sql(f"  SELECT {SESSION_ID}, s.id, cs.class_id, cs.section_id, 'yes', 0")
    sql(f"  FROM students s")
    sql(f"  JOIN classes c   ON LOWER(c.class)   = LOWER({esc(cls)})")
    sql(f"  JOIN sections sec ON LOWER(sec.section) = LOWER({esc(sec)})")
    sql(f"  JOIN class_sections cs ON cs.class_id=c.id AND cs.section_id=sec.id")
    sql(f"  WHERE s.note={esc(BATCH_TAG)} AND s.admission_no IN (")
    sql(f"    {', '.join(esc(a) for a in cls_adm_nos)}")
    sql(f"  )")
    sql(f"  AND NOT EXISTS (")
    sql(f"    SELECT 1 FROM student_session ss2")
    sql(f"    WHERE ss2.session_id={SESSION_ID} AND ss2.student_id=s.id")
    sql(f"  );")
    sql("")

# ─── 7. USER ACCOUNTS ────────────────────────────────────────────────────────
sql("-- ── 7. Student user accounts (username = admission_no) ───────────────")
sql("INSERT INTO users (user_id, username, password, childs, role, lang_id, currency_id, verification_code, is_active)")
sql(f"SELECT s.id, s.admission_no, 'Welcome@123', '', 'student', {LANG_ID}, {CURRENCY_ID}, '', 'yes'")
sql(f"FROM students s WHERE s.note={esc(BATCH_TAG)}")
sql("AND NOT EXISTS (SELECT 1 FROM users u WHERE u.user_id=s.id AND u.role='student');")
sql("")

sql("-- Parent user accounts (username = par{student_id})")
sql("INSERT INTO users (user_id, username, password, childs, role, lang_id, currency_id, verification_code, is_active)")
sql(f"SELECT s.id, CONCAT('par', s.id), 'Welcome@123', s.id, 'parent', {LANG_ID}, {CURRENCY_ID}, '', 'yes'")
sql(f"FROM students s WHERE s.note={esc(BATCH_TAG)}")
sql("AND NOT EXISTS (SELECT 1 FROM users u WHERE u.username=CONCAT('par', s.id) AND u.role='parent');")
sql("")

sql("-- Set parent_id on students")
sql(f"UPDATE students s")
sql(f"JOIN users u ON u.username=CONCAT('par', s.id) AND u.role='parent'")
sql(f"SET s.parent_id=u.id")
sql(f"WHERE s.note={esc(BATCH_TAG)};")
sql("")

sql("SET FOREIGN_KEY_CHECKS = 1;")
sql("-- ── End of maasc_reimport_2026.sql ────────────────────────────────────")

# ─────────────────────────────────────────────────────────────────────────────
# WRITE SQL FILE
# ─────────────────────────────────────────────────────────────────────────────
with open(OUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

print(f"\nSQL written: {OUT_SQL}")
print(f"  Unique class-section pairs: {len(class_section_pairs)}")

PEM  = "/Volumes/WORK/aws ec2 connect/minerva_prod.pem"
HOST = "ec2-user@13.234.255.106"
print(f"\nTo deploy manually:")
print(f"  scp -i '{PEM}' {OUT_SQL} {HOST}:/tmp/maasc_reimport_2026.sql")
print(f"  ssh -i '{PEM}' {HOST} \"mysql -u meenakshi -p'3gw4*86*!zno-EMO' maasc < /tmp/maasc_reimport_2026.sql\"")

if '--deploy' in sys.argv:
    print("\nDeploying to EC2...")
    r1 = subprocess.run(
        ['scp', '-i', PEM, '-o', 'StrictHostKeyChecking=no',
         OUT_SQL, f'{HOST}:/tmp/maasc_reimport_2026.sql'],
        capture_output=True, text=True)
    if r1.returncode != 0:
        print("SCP failed:", r1.stderr)
        sys.exit(1)
    print("SCP OK. Running SQL...")
    r2 = subprocess.run(
        ['ssh', '-i', PEM, '-o', 'StrictHostKeyChecking=no', HOST,
         "mysql -u meenakshi -p'3gw4*86*!zno-EMO' maasc < /tmp/maasc_reimport_2026.sql"],
        capture_output=True, text=True)
    if r2.returncode != 0:
        print("MySQL failed:", r2.stderr)
        sys.exit(1)
    print("Done! MySQL output:", r2.stdout or "(no output = success)")
