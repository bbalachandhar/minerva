from io import BytesIO
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path('/Applications/XAMPP/xamppfiles/htdocs/minerva')
IN_JSON = ROOT / 'docs' / 'review_comments_jan2026_analysis.json'
OUT_XLSX = ROOT / 'Jan2026_review_comments_leadership_summary.xlsx'


def style_header(ws, row_idx):
	fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
	for cell in ws[row_idx]:
		if cell.value is None:
			continue
		cell.font = Font(color='FFFFFF', bold=True)
		cell.fill = fill
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def autosize(ws, min_width=12, max_width=50):
	widths = {}
	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
		for cell in row:
			if cell.value is None:
				continue
			widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)) + 2)
	for col_idx, width in widths.items():
		ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max(min_width, min(max_width, width))


def add_title(ws, row_idx, title):
	ws.cell(row=row_idx, column=1, value=title)
	ws.cell(row=row_idx, column=1).font = Font(bold=True, color='1F4E78')


def main():
	payload = json.loads(IN_JSON.read_text(encoding='utf-8'))
	combined = payload['combined']
	sheets = payload['sheet_results']

	wb = Workbook()
	ws = wb.active
	ws.title = 'Leadership_Summary'

	r = 1
	add_title(ws, r, "Executive Snapshot")
	r += 1
	ws.append(['Metric', 'Value'])
	style_header(ws, r)
	ws.append(['Workbook', Path(payload['workbook']).name])
	ws.append(['Total Comments', combined['total_rows']])
	ws.append(['Critical Comments', combined['critical_count']])
	ws.append(['Non-Critical Comments', None])
	ws.append(['Critical %', None])
	ws.append(['Non-Critical %', None])
	ws.append(['Legacy Total', sheets['Legacy']['total_rows']])
	ws.append(['Legacy Critical', sheets['Legacy']['critical_count']])
	ws.append(['Legacy Critical %', None])
	ws.append(['Asset west Total', sheets['Asset west']['total_rows']])
	ws.append(['Asset west Critical', sheets['Asset west']['critical_count']])
	ws.append(['Asset west Critical %', None])

	# Formula-driven KPIs for future-proof manual updates.
	ws['B6'] = '=B4-B5'
	ws['B7'] = '=IFERROR(B5/B4,0)'
	ws['B8'] = '=IFERROR(B6/B4,0)'
	ws['B11'] = '=IFERROR(B10/B9,0)'
	ws['B14'] = '=IFERROR(B13/B12,0)'
	for cell in ('B7', 'B8', 'B11', 'B14'):
		ws[cell].number_format = '0.00%'

	r = ws.max_row + 2
	add_title(ws, r, "Top Risk Drivers")
	r += 1
	ws.append(['Type', 'Rank', 'Label', 'Count'])
	style_header(ws, r)
	for idx, item in enumerate(combined['top_categories'][:5], start=1):
		ws.append(['Category', idx, item['label'], item['count']])
	for idx, item in enumerate(combined['top_sub_categories'][:5], start=1):
		ws.append(['Sub-category', idx, item['label'], item['count']])

	r = ws.max_row + 2
	add_title(ws, r, "Leadership Actions (This Week)")
	r += 1
	ws.append(['Priority', 'Action', 'Expected Outcome'])
	style_header(ws, r)
	ws.append(['P1', 'Close top hotspot critical items first (Y86, J69, F55, D29, J08)', 'Visible critical reduction in 48 hours'])
	ws.append(['P1', 'Run daily dependency unblock huddle with Controller + RM + reviewers', 'Faster clearance of blocked line items'])
	ws.append(['P2', 'Operate dual lanes: Asset west volume + Legacy critical-rate correction', 'Balanced reduction of risk and queue load'])
	ws.append(['P2', 'Standardize category naming variants across tracker', 'Cleaner reporting and trend reliability'])

	ws2 = wb.create_sheet('Property_Hotspots')
	add_title(ws2, 1, 'Top Critical Hotspots')
	ws2.append(['Rank', 'Property Code', 'Property Name', 'Critical Count'])
	style_header(ws2, 2)
	for idx, item in enumerate(combined['critical_hotspots'][:15], start=1):
		ws2.append([idx, item['property_code'], item['property_name'], item['count']])

	ws3 = wb.create_sheet('Instructions')
	add_title(ws3, 1, 'How To Use This File')
	ws3.append(['Step', 'Instruction'])
	style_header(ws3, 2)
	ws3.append(['1', 'Update source data in review workbook and regenerate JSON via tools/review_comments_analysis.py'])
	ws3.append(['2', 'Run tools/build_leadership_summary_xlsx.py to refresh this report'])
	ws3.append(['3', 'If editing manually, update only numeric input cells: B4, B5, B9, B10, B12, B13 in Leadership_Summary'])
	ws3.append(['4', 'Do not overwrite formula cells: B6, B7, B8, B11, B14'])
	ws3.append(['5', 'Critical and Non-Critical percentages auto-calculate from formulas'])
	ws3.append(['6', 'Use Property_Hotspots tab for leadership escalation and ownership assignment'])

	add_title(ws3, 10, 'Formula Reference')
	ws3.append(['Cell', 'Formula / Purpose'])
	style_header(ws3, 11)
	ws3.append(['Leadership_Summary!B6', 'Non-Critical Comments = Total - Critical'])
	ws3.append(['Leadership_Summary!B7', 'Critical % = Critical / Total'])
	ws3.append(['Leadership_Summary!B8', 'Non-Critical % = Non-Critical / Total'])
	ws3.append(['Leadership_Summary!B11', 'Legacy Critical % = Legacy Critical / Legacy Total'])
	ws3.append(['Leadership_Summary!B14', 'Asset west Critical % = Asset west Critical / Asset west Total'])

	autosize(ws)
	autosize(ws2)
	autosize(ws3)

	buffer = BytesIO()
	wb.save(buffer)
	OUT_XLSX.write_bytes(buffer.getvalue())
	print(str(OUT_XLSX))


if __name__ == '__main__':
	main()
